"""
demm_tiempo_real.py
===================
Analisis NVH en tiempo real — DEMM
Dos ventanas con estilo identico al explorador_dataset.

Ventana 1 — Panel Operador: datos operador, modelos, indicadores BUENO/MALO
Ventana 2 — Tabla de Turno: historial de ciclos + exportar Excel

USO:
  py demm_tiempo_real.py
REQUISITOS:
  pip install sounddevice numpy scipy scikit-learn openpyxl matplotlib
"""

import os, sys, time, pickle, threading, collections, warnings, datetime

def _resource_path(relpath):
    """
    Resuelve rutas de recursos tanto en modo .py como en .exe (PyInstaller).
    Busca en este orden:
      1. Carpeta del ejecutable (dist/MatIA/) — para modelos .pkl sueltos
      2. sys._MEIPASS — para recursos embebidos dentro del exe (logo, ico)
      3. Carpeta del script .py — para desarrollo
    """
    # Carpeta del ejecutable (cuando corre como .exe empaquetado)
    if getattr(sys, 'frozen', False):
        exe_dir = os.path.dirname(sys.executable)
        candidate = os.path.join(exe_dir, relpath)
        if os.path.exists(candidate):
            return candidate
        # Recursos embebidos en el exe
        meipass = getattr(sys, "_MEIPASS", exe_dir)
        return os.path.join(meipass, relpath)
    # Modo desarrollo (.py directo)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), relpath)
import tkinter as tk
from tkinter import filedialog, ttk, messagebox

import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.animation import FuncAnimation

try:
    import sounddevice as sd
    SD_OK = True
except ImportError:
    SD_OK = False

warnings.filterwarnings("ignore")

# ── Parametros señal ────────────────────────────────────────────────────────
FS     = 48000
RPM    = 872.0
F_ROT  = RPM / 60.0          # 7.333 Hz
T_ROT  = 1.0 / F_ROT
M_ROT  = int(T_ROT * FS)
BUF_N  = FS * 6

PINONES = {
    "PIMA":         {"dientes": 26, "gmf": 26 * F_ROT},
    "ARBOL_SEC_14": {"dientes": 14, "gmf": 14 * F_ROT},
    "ARBOL_SEC_15": {"dientes": 15, "gmf": 15 * F_ROT},
}

# ── Detección de engrane por RMS (igual que grabador automático) ────────────
DURACION_ESTAB    = 0.80    # segundos continuos sobre umbral → engrane confirmado
                            # Mantenido en 0.80s — el empuje dura ~1.69s pero
                            # el discriminador real son los filtros CV + periodicidad
MUESTRAS_ESTAB    = int(FS * DURACION_ESTAB)
DURACION_SILENCIO = 0.08    # segundos bajo umbral → resetear contador
MUESTRAS_SILENCIO = int(FS * DURACION_SILENCIO)
RMS_FONDO         = 0.00857
FACTOR_UMBRAL     = 1.5
UMBRAL_RMS_FIJO   = round(RMS_FONDO * FACTOR_UMBRAL, 6)
MAX_CV            = 1.2      # CV máximo deslizante (reducido de 1.8)
MIN_CICLOS_PERIODO = 4       # giros periódicos mínimos para confirmar engrane real
DURACION_CICLO    = 2.5
CALIBRACION_SEG   = 3.0
# ── Buffer pre-engrane ───────────────────────────────────────────────────────
DURACION_PREENGRANE  = DURACION_ESTAB
MUESTRAS_PREENGRANE  = int(FS * DURACION_PREENGRANE)

# ── Paleta identica al explorador_dataset ──────────────────────────────────
C_BG           = "#f0f2f5"   # fondo principal — gris muy claro
C_SURFACE      = "#ffffff"   # superficies (headers, paneles)
C_SURFACE2     = "#e8eaed"   # superficies secundarias (entradas, botones)
C_BORDER       = "#c8cdd8"   # bordes visibles
C_BORDER2      = "#b0b6c4"   # bordes secundarios
C_TEXT         = "#1a1d27"   # texto principal — casi negro
C_TEXT_SUB     = "#4a5068"   # texto secundario — gris medio
C_TEXT_DIM     = "#8a90a8"   # texto atenuado — gris claro
C_ACENTO       = "#1a5fa8"   # azul corporativo (más oscuro para contraste sobre blanco)
C_BUENO        = "#16a34a"   # verde diagnóstico
C_MALO         = "#dc2626"   # rojo diagnóstico
C_REVISAR      = "#d97706"   # ámbar diagnóstico
C_INVALIDO_VAL = "#64748b"   # gris neutro
C_MONO         = "Consolas"

FEATURES = ["K_ret","CF_p99_ret","K_emp","CF_p99_emp","rms_ret","rms_emp"]

# ── Logo empresa ─────────────────────────────────────────────────────────────
_LOGO_TK   = None   # PhotoImage — se carga una sola vez
_MATIA_TK = None   # PhotoImage logo MatIA

def _cargar_logo():
    """
    Carga logo_empresa.png desde la misma carpeta del script.
    Redimensiona a 110x40 px máximo preservando proporciones.
    Retorna un PhotoImage listo para tk.Label, o None si no existe.
    """
    global _LOGO_TK
    if _LOGO_TK is not None:
        return _LOGO_TK
    try:
        from PIL import Image, ImageTk
        ruta = _resource_path("logo_empresa.png")
        if not os.path.exists(ruta):
            return None
        img = Image.open(ruta).convert("RGBA")
        img.thumbnail((110, 40), Image.LANCZOS)
        _LOGO_TK = ImageTk.PhotoImage(img)
        return _LOGO_TK
    except Exception:
        return None


# Cache de imágenes de piñones
_IMG_PIMA_TK    = None
_IMG_ARBOL_TK   = None

def _cargar_img_pinon(tipo, ancho=180, alto=110):
    """
    Carga la imagen del piñón desde la carpeta HORSE.
    tipo: 'PIMA' o 'ARBOL'
    Retorna PhotoImage o None si no existe.
    """
    global _IMG_PIMA_TK, _IMG_ARBOL_TK
    if tipo == "PIMA"  and _IMG_PIMA_TK  is not None: return _IMG_PIMA_TK
    if tipo == "ARBOL" and _IMG_ARBOL_TK is not None: return _IMG_ARBOL_TK
    try:
        from PIL import Image, ImageTk
        nombre = "img_pima.png" if tipo == "PIMA" else "img_arbol_sec.png"
        ruta = _resource_path(nombre)
        if not os.path.exists(ruta):
            return None
        img = Image.open(ruta).convert("RGBA")
        img.thumbnail((ancho, alto), Image.LANCZOS)
        tk_img = ImageTk.PhotoImage(img)
        if tipo == "PIMA":  _IMG_PIMA_TK  = tk_img
        else:               _IMG_ARBOL_TK = tk_img
        return tk_img
    except Exception:
        return None


def _cargar_logo_matia():
    """
    Carga matia_logo.png desde la misma carpeta del script.
    Si no existe lo genera en tiempo de ejecución con PIL.
    Retorna PhotoImage listo para tk.Label.
    """
    global _MATIA_TK
    if _MATIA_TK is not None:
        return _MATIA_TK
    try:
        from PIL import Image, ImageTk, ImageDraw, ImageFont
        import math as _math

        ruta = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            "matia_logo.png")

        if os.path.exists(ruta):
            img = Image.open(ruta).convert("RGBA")
        else:
            # Generar el logo si no existe el PNG
            SCALE = 4
            W, H  = 240*SCALE, 60*SCALE
            img   = Image.new("RGBA", (W, H), (0, 0, 0, 0))
            dr    = ImageDraw.Draw(img)
            C_DARK = (26,29,39,255); C_BLUE=(26,95,168,255); C_MID=(74,80,104,255)
            gx,gy = 34*SCALE,30*SCALE
            r_ext,r_int = 22*SCALE,14*SCALE; n_d=8
            pts=[]
            for i in range(n_d*2):
                ang=_math.pi*i/n_d-_math.pi/2
                r=r_ext if i%2==0 else r_int
                pts.append((gx+r*_math.cos(ang),gy+r*_math.sin(ang)))
            dr.polygon(pts,fill=C_DARK); dr.polygon(pts,outline=C_BLUE,width=round(2.2*SCALE/3))
            for i in range(n_d):
                a0=_math.pi*(2*i)/n_d-_math.pi/2; a2=_math.pi*(2*i+2)/n_d-_math.pi/2
                dr.line([gx+r_ext*_math.cos(a0),gy+r_ext*_math.sin(a0),
                         gx+r_ext*_math.cos(a2),gy+r_ext*_math.sin(a2)],
                        fill=C_BLUE,width=round(2*SCALE/3))
            ri=9*SCALE
            dr.ellipse([gx-ri,gy-ri,gx+ri,gy+ri],fill=C_DARK,outline=C_BLUE,width=round(1.5*SCALE/3))
            nc=5*SCALE//3; dr.ellipse([gx-nc,gy-nc,gx+nc,gy+nc],fill=C_BLUE)
            for ad in [90,0,270,180]:
                a=_math.radians(ad)
                x1=gx+nc*_math.cos(a); y1=gy+nc*_math.sin(a)
                x2=gx+(ri-round(2*SCALE/3))*_math.cos(a); y2=gy+(ri-round(2*SCALE/3))*_math.sin(a)
                dr.line([x1,y1,x2,y2],fill=C_MID,width=max(1,round(1.8*SCALE/3)))
                sr=max(2,round(3.5*SCALE/3)); dr.ellipse([x2-sr,y2-sr,x2+sr,y2+sr],fill=C_BLUE)
            def lf(ps,sz):
                for p in ps:
                    if os.path.exists(p):
                        try: return ImageFont.truetype(p,sz)
                        except: pass
                return ImageFont.load_default()
            sz=round(26*SCALE/3); szt=round(9*SCALE/3)
            fb=lf(["/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
                   "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf"],sz)
            fr=lf(["/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
                   "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf"],sz)
            fm=lf(["/usr/share/fonts/truetype/dejavu/DejaVuSansMono.ttf",
                   "/usr/share/fonts/truetype/liberation/LiberationMono-Regular.ttf"],szt)
            tx=round(68*SCALE/3); ty=round(4*SCALE/3)
            dr.text((tx,ty),"Mat",font=fb,fill=C_DARK)
            bb_m=dr.textbbox((tx,ty),"Mat",font=fb)
            x_ia=bb_m[2]+round(3*SCALE/3)
            dr.text((x_ia,ty),"IA",font=fr,fill=C_BLUE)
            bb_ia=dr.textbbox((x_ia,ty),"IA",font=fr)
            ly=bb_ia[3]+round(4*SCALE/3)
            dr.rectangle([tx,ly,bb_ia[2],ly+round(3*SCALE/3)],fill=C_BLUE)
            dr.text((tx,ly+round(7*SCALE/3)),"NVH  ·  DEMM",font=fm,fill=C_MID)
            img = img.resize((220,56),Image.LANCZOS)
            try: img.save(ruta,"PNG")
            except Exception: pass

        img.thumbnail((220, 56), Image.LANCZOS)
        _MATIA_TK = ImageTk.PhotoImage(img)
        return _MATIA_TK
    except Exception:
        return None

def _hacer_logo_matia(parent, bg=None):
    """Coloca el logo MatIA como Label en el widget padre."""
    bg  = bg or C_SURFACE
    img = _cargar_logo_matia()
    if img:
        lbl = tk.Label(parent, image=img, bg=bg)
        lbl.image = img   # mantener referencia
        return lbl
    # Fallback texto si PIL no disponible
    return tk.Label(parent, text="MatIA", bg=bg, fg="#1a5fa8",
                    font=("Arial", 11, "bold"))

# Rutas de modelos por defecto
# Buscar modelos: junto al exe → carpeta HORSE → carpeta del script
_BASE = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop", "HORSE")
def _ruta_modelo(nombre):
    # 1. Junto al ejecutable (instalación normal con Setup)
    if getattr(sys, 'frozen', False):
        exe_dir = os.path.dirname(sys.executable)
        p = os.path.join(exe_dir, nombre)
        if os.path.exists(p): return p
    # 2. Carpeta HORSE en el escritorio (desarrollo)
    p = os.path.join(_BASE, nombre)
    if os.path.exists(p): return p
    # 3. Carpeta del script
    p = os.path.join(os.path.dirname(os.path.abspath(__file__)), nombre)
    if os.path.exists(p): return p
    # Fallback — devuelve ruta junto al exe aunque no exista
    if getattr(sys, 'frozen', False):
        return os.path.join(os.path.dirname(sys.executable), nombre)
    return os.path.join(_BASE, nombre)
RUTA_MODELO_ID   = _ruta_modelo("modelo_identificador_20260401_1251.pkl")
RUTA_MODELO_PIMA = _ruta_modelo("modelo_nvh_PIMA_20260317_0920.pkl")
RUTA_MODELO_AS   = _ruta_modelo("modelo_nvh_ARB_14_20260319_0950.pkl")
RUTA_UMB_AS14    = _ruta_modelo("umbrales_ARBOL_SEC_14.pkl")
RUTA_UMB_PIMA    = _ruta_modelo("umbrales_PIMA.pkl")

UMBRALES = {
    "PIMA":         {"K_ret": 4.04, "K_emp": 3.89, "CF_ret": 3.27, "CF_emp": 3.34},
    "ARBOL_SEC_14": {"K_ret": 5.23, "K_emp": 4.95, "CF_ret": 3.64, "CF_emp": 3.61},
    "ARBOL_SEC_15": {"K_ret": 5.23, "K_emp": 4.95, "CF_ret": 3.64, "CF_emp": 3.61},
}

# ═══════════════════════════════════════════════════════════════════════════
# SEGMENTACION (identica a etiquetar_dataset / explorador)
# ═══════════════════════════════════════════════════════════════════════════

def _detectar_separador(s):
    n = len(s); s = s.astype(np.float64)
    i0 = int(n*0.15); i1 = int(n*0.85); zona = s[i0:i1]; ng = len(zona)//M_ROT
    if ng < 4:
        env = np.sqrt(np.convolve(s**2, np.ones(max(1,int(0.20*FS)))/max(1,int(0.20*FS)), mode="same"))
        return int(np.argmax(env[i0:i1]))+i0
    rg  = np.array([float(np.sqrt(np.mean(zona[i*M_ROT:(i+1)*M_ROT]**2))) for i in range(ng)])
    ref = float(np.median(rg[:max(2,ng//3)])); ref = ref if ref>1e-9 else float(np.median(rg))+1e-9
    for f in [4.0, 3.0, 2.5]:
        for i, r in enumerate(rg):
            if r > ref*f: return i0+i*M_ROT
    return i0+int(np.argmax(rg))*M_ROT

def _detectar_freno(z):
    n = len(z)
    if n < int(FS*0.15): return n
    s = z.astype(np.float64); ng = n//M_ROT
    if ng < 6: return int(n*0.92)
    rg = np.array([float(np.sqrt(np.mean(s[i*M_ROT:(i+1)*M_ROT]**2))) for i in range(ng)])
    MG = max(7, int(0.35*FS/M_ROT)); i1 = min(MG+5, ng-1)
    if i1 <= MG or MG >= ng: return int(n*0.92)
    ref = float(np.median(rg[MG:i1]))
    if ref < 1e-9: return int(n*0.92)
    for i in range(MG, ng):
        if rg[i] > ref*2.5: return max(0, i*M_ROT-M_ROT//2)
    return int(n*0.92)

def _detectar_estabilizacion(z):
    s = z.astype(np.float64); n = len(s); ng = n//M_ROT
    if ng < 3: return 0
    rg  = np.array([float(np.sqrt(np.mean(s[i*M_ROT:(i+1)*M_ROT]**2))) for i in range(ng)])
    ref = float(np.percentile(rg[ng//2:], 20))
    if ref < 1e-9: ref = float(np.median(rg))
    if ref < 1e-9: return 0
    mn = ref*0.35; mx = ref*1.80; GM = 5; CN = 2; ini = min(GM, ng-CN-1)
    for i in range(ini, ng-CN+1):
        v = rg[i:i+CN]
        if not all(mn <= r <= mx for r in v): continue
        if np.max(v)/(np.min(v)+1e-12) > 1.8: continue
        return max(0, i*M_ROT)
    return min(GM*M_ROT, n//3)

def _kg(z):
    z = z.astype(np.float64); mu = np.mean(z)
    return float(np.mean((z-mu)**4)/(np.mean((z-mu)**2)**2+1e-12))

def _cf(z):
    z = z.astype(np.float64); rms = float(np.sqrt(np.mean(z**2)))
    return float(np.percentile(np.abs(z), 99.5))/(rms+1e-12)

def _tramo_central_rt(z, tramo=int(0.5 * FS)):
    """Extrae el tramo central de una zona para FFT en tiempo real."""
    if len(z) <= tramo:
        return z
    ini = max(0, len(z) // 2 - tramo // 2)
    return z[ini:ini + tramo]


def _espectro_rt(z):
    """FFT en órdenes para tiempo real. Retorna (ordenes, dB)."""
    s = z.astype(np.float64)
    nn = len(s)
    if nn < 256:
        return np.array([0.0]), np.array([0.0])
    vh  = np.hanning(nn)
    mag = np.abs(np.fft.rfft(s * vh)) * 2 / nn
    fq  = np.fft.rfftfreq(nn, d=1.0 / FS)
    db  = 20 * np.log10(mag + 1e-12)
    return fq / F_ROT, db


def analizar_ciclo(senal):
    n = len(senal)
    if n < int(FS*1.5): return None
    try:
        mg = int(FS*0.02); xs = _detectar_separador(senal)
        ei = int(n*0.05);  ef = max(0, xs-mg)
        ri = min(n, xs+mg)
        zrp = senal[ri:int(n*0.95)]; xf = _detectar_freno(zrp); rf = ri+xf
        zrb = senal[ri:rf];           xe = _detectar_estabilizacion(zrb)
        ze  = senal[ei:ef].astype(np.float64)
        zr  = zrb[xe:].astype(np.float64)
        if len(zr) < M_ROT*3 or len(ze) < M_ROT*2: return None
        zrl = zr[:-M_ROT]
        zeu = ze[:-2*M_ROT] if len(ze) > 2*M_ROT*2 else ze

        resultado = {
            "K_ret":      round(_kg(zrl), 3),
            "CF_p99_ret": round(_cf(zrl), 3),
            "K_emp":      round(_kg(zeu), 3),
            "CF_p99_emp": round(_cf(zeu), 3),
            "rms_ret":    round(float(np.sqrt(np.mean(zr**2))), 5),
            "rms_emp":    round(float(np.sqrt(np.mean(ze**2))), 5),
            "dur_ret":    round(len(zr)/FS, 3),
            "consumir":   rf+int(FS*0.2),
            "seg_ret":    zrl.copy(),   # segmento retroceso para visualización
            "seg_emp":    zeu.copy(),   # segmento empuje para visualización
        }
        # Espectro FFT opcional — si falla no afecta el análisis NVH
        try:
            ords_r, db_r = _espectro_rt(_tramo_central_rt(zrl))
            ords_e, db_e = _espectro_rt(_tramo_central_rt(zeu))
            if ords_r is not None:
                resultado["ords_ret"] = ords_r
                resultado["db_ret"]   = db_r
            if ords_e is not None:
                resultado["ords_emp"] = ords_e
                resultado["db_emp"]   = db_e
        except Exception:
            pass  # sireneo no disponible en este ciclo, NVH continúa igual
        return resultado
    except Exception:
        return None

# ═══════════════════════════════════════════════════════════════════════════
# IDENTIFICACION Y CLASIFICACION
# ═══════════════════════════════════════════════════════════════════════════

def _feats_id(senal):
    """
    Extrae las mismas 18 features que usa el modelo identificador entrenado.
    Debe ser idéntica a _extraer_features_csv() del entrenador.
    """
    try:
        s  = senal.astype(np.float64)
        n  = len(s)
        seg = s[n//4 : 3*n//4]
        ns  = len(seg)
        if ns < 256:
            return {}

        ventana = np.hanning(ns)
        fm  = np.abs(np.fft.rfft(seg * ventana)) * 2 / ns
        fr  = np.fft.rfftfreq(ns, d=1.0/FS)
        fdb = 20 * np.log10(fm + 1e-12)

        def energia_banda(fc, bw=15):
            m = (fr >= fc-bw) & (fr <= fc+bw)
            return float(np.sum(fm[m]**2)) if np.any(m) else 0.0

        def pico_banda(fc, bw=15):
            m = (fr >= fc-bw) & (fr <= fc+bw)
            return float(np.max(fm[m])) if np.any(m) else 0.0

        rms    = float(np.sqrt(np.mean(seg**2)))
        mu     = np.mean(seg); sigma2 = np.var(seg)
        kurt   = float(np.mean((seg-mu)**4) / (sigma2**2 + 1e-12)) if sigma2 > 0 else 3.0

        n_vent = ns // M_ROT
        rms_v  = np.array([float(np.sqrt(np.mean(seg[i*M_ROT:(i+1)*M_ROT]**2)))
                            for i in range(n_vent)]) if n_vent > 1 else np.array([rms])
        mean_v = float(np.mean(rms_v))
        cv_rms = float(np.std(rms_v) / mean_v) if mean_v > 1e-9 else 0.0

        fondo_espectro = float(np.percentile(fdb, 20))
        densidad_ruido = float(np.std(fdb))
        umb_db         = np.percentile(fdb, 85)
        n_frec_umbral  = int(np.sum(fdb > umb_db))
        nivel_max      = float(np.max(fdb))
        freq_pico1     = float(fr[np.argmax(fm)])
        freq_centroide = float(np.sum(fr * fm**2) / (np.sum(fm**2) + 1e-12))

        e_total = float(np.sum(fm**2)) + 1e-12
        gmf_p   = PINONES["PIMA"]["gmf"]
        gmf_14  = PINONES["ARBOL_SEC_14"]["gmf"]

        e_gmf_pima   = energia_banda(gmf_p)
        e_gmf_arb14  = energia_banda(gmf_14)
        e_2gmf_pima  = energia_banda(gmf_p  * 2)
        e_2gmf_arb14 = energia_banda(gmf_14 * 2)

        gmf_key        = gmf_14   # usar AS14 como referencia para ratios generales
        ratio_gmf      = float(energia_banda(gmf_key) / e_total)
        ratio_2gmf     = float(energia_banda(gmf_key*2) / e_total)
        ratio_3gmf     = float(energia_banda(gmf_key*3) / e_total)
        ratio_ruido_gmf= float(energia_banda(gmf_key) / (float(np.mean(fm**2)) + 1e-12))
        ratio_hilo     = float(pico_banda(gmf_key) / (float(np.mean(fm)) + 1e-12))

        return {
            "rms": rms, "kurt": kurt, "cv_rms": cv_rms,
            "fondo_espectro": fondo_espectro, "densidad_ruido": densidad_ruido,
            "n_frec_umbral": n_frec_umbral, "nivel_max": nivel_max,
            "ratio_gmf": ratio_gmf, "ratio_2gmf": ratio_2gmf,
            "ratio_3gmf": ratio_3gmf, "ratio_ruido_gmf": ratio_ruido_gmf,
            "freq_centroide": freq_centroide, "ratio_hilo": ratio_hilo,
            "e_gmf_pima":   e_gmf_pima,  "e_gmf_arb14":  e_gmf_arb14,
            "e_2gmf_pima":  e_2gmf_pima, "e_2gmf_arb14": e_2gmf_arb14,
            "freq_pico1": freq_pico1,
        }
    except Exception:
        return {}

def identificar_pinon(senal, mid, hist_pinon=None):
    """
    Identifica el piñón usando el modelo ML.
    Filtro de confianza mínima: si conf < 0.65 y hay historial,
    usa el piñón más frecuente reciente en vez de aceptar una
    predicción poco confiable.
    No aplica filtro de consistencia temporal para permitir
    cambios de piñón en cualquier momento.
    """
    CONF_MIN = 0.65

    if mid is None:
        return "DESCONOCIDO", 0.0
    try:
        pipe  = mid["pipeline"]; le = mid["label_enc"]; cols = mid["feature_cols"]
        f     = _feats_id(senal)
        if not f:
            return "DESCONOCIDO", 0.0
        x     = np.array([[f.get(c, 0.0) for c in cols]])
        clase = le.inverse_transform([pipe.predict(x)[0]])[0]
        conf  = float(np.max(pipe.predict_proba(x)[0]))

        # Normalizar clase
        if "ARB_14" in clase.upper():
            pk = "ARBOL_SEC_14"
        elif "ARB_15" in clase.upper():
            pk = "ARBOL_SEC_15"
        else:
            pk = "PIMA"

        # Filtro de confianza mínima — si el modelo no está seguro
        # usa el piñón más frecuente de los últimos ciclos
        if conf < CONF_MIN and hist_pinon is not None and len(hist_pinon) >= 2:
            from collections import Counter
            pk = Counter(hist_pinon).most_common(1)[0][0]

        # Actualizar historial
        if hist_pinon is not None:
            hist_pinon.append(pk)

        return pk, conf
    except Exception:
        return "DESCONOCIDO", 0.0

def clasificar_nvh(feats, pinon_key, modelos):
    m = modelos.get(pinon_key)
    if m is None: return "SIN MODELO", 0.0
    try:
        pipe = m["pipeline"]; cols = m.get("feature_cols", FEATURES)
        x    = np.array([[feats[c] for c in cols]])
        pred = pipe.predict(x)[0]; probs = pipe.predict_proba(x)[0]
        return ("MALO" if pred==1 else "BUENO"), float(np.max(probs))
    except: return "ERROR", 0.0

def tag_por_umbrales(ciclo):
    """Determina el tag de color segun umbrales K y CF."""
    pinon = ciclo.get("pinon", "ARBOL_SEC_14")
    umb   = UMBRALES.get(pinon, UMBRALES["ARBOL_SEC_14"])
    ku_r  = umb["K_ret"]; ku_e = umb["K_emp"]
    cfu_r = umb["CF_ret"]; cfu_e = umb["CF_emp"]
    kr=ciclo.get("K_ret",0); cr=ciclo.get("CF_ret",0)
    ke=ciclo.get("K_emp",0); ce=ciclo.get("CF_emp",0)
    malo_ret = (kr>ku_r and cr>cfu_r); malo_emp = (ke>ku_e and ce>cfu_e)
    sosp_ret = (kr>ku_r) != (cr>cfu_r); sosp_emp = (ke>ku_e) != (ce>cfu_e)
    if malo_ret or malo_emp:   return "MALO"
    if sosp_ret or sosp_emp:   return "REVISAR"
    return "BUENO"

def detectar_golpe(feats, pinon_key, modelos_nvh=None):
    """
    Detecta GOLPE con lógica AND combinada:
      1. Umbrales físicos: K > umbral AND CF.p99 > umbral (en ret o emp)
      2. Modelo NVH:       clasifica como MALO
    Ambas condiciones deben cumplirse simultáneamente.
    Si no hay modelo NVH cargado, usa solo los umbrales.
    Retorna (bool_golpe, detalle_str).
    """
    umb      = UMBRALES.get(pinon_key, UMBRALES["ARBOL_SEC_14"])
    ku_r     = umb["K_ret"];  ku_e  = umb["K_emp"]
    cfu_r    = umb["CF_ret"]; cfu_e = umb["CF_emp"]
    kr       = feats.get("K_ret", 0);  cr = feats.get("CF_p99_ret", 0)
    ke       = feats.get("K_emp", 0);  ce = feats.get("CF_p99_emp", 0)
    malo_ret = (kr > ku_r and cr > cfu_r)
    malo_emp = (ke > ku_e and ce > cfu_e)
    umbral_supera = malo_ret or malo_emp
    zona = ("RET+EMP" if malo_ret and malo_emp
            else "RET" if malo_ret else "EMP" if malo_emp else "—")

    # Si no supera umbrales físicos → no hay golpe sin importar el modelo
    if not umbral_supera:
        return False, "—"

    # Si hay modelo NVH, exigir que también diga MALO (lógica AND)
    if modelos_nvh:
        etq_nvh, conf_nvh = clasificar_nvh(feats, pinon_key, modelos_nvh)
        if etq_nvh != "MALO":
            return False, "—"   # umbrales superados pero modelo dice BUENO → no es golpe

    return True, zona


def detectar_sireneo(feats, pinon_key, umbral_espectro, tol_ordenes=2.0):
    """
    Detecta SIRENEO: pico en 1×GMF, 2×GMF o 3×GMF que supera el umbral
    estadístico del dataset en ese orden.
    Retorna (bool_sireneo, lista de armónicos afectados).
    """
    try:
        ud = umbral_espectro.get(pinon_key)
        if not ud:
            return False, []

        gmf_hz = PINONES.get(pinon_key, {}).get("gmf", 0)
        if gmf_hz == 0:
            return False, []
        orden_gmf = gmf_hz / F_ROT
        grilla    = ud.get("ordenes")

        armónicos_detectados = []

        for flanco, key_ords, key_db in [
            ("ret", "ords_ret", "db_ret"),
            ("emp", "ords_emp", "db_emp"),
        ]:
            ud_flanco = ud.get(flanco)
            if not ud_flanco or grilla is None:
                continue
            ords = feats.get(key_ords)
            dbs  = feats.get(key_db)
            if ords is None or dbs is None or len(ords) < 2:
                continue

            for k in [1, 2, 3]:
                o_k = orden_gmf * k
                banda_sig = (ords >= o_k - tol_ordenes) & (ords <= o_k + tol_ordenes)
                if not np.any(banda_sig):
                    continue
                val_peak = float(np.max(dbs[banda_sig]))
                umb_val  = float(np.interp(o_k, grilla, ud_flanco["umbral"]))
                if val_peak > umb_val:
                    armónicos_detectados.append((k, flanco, round(val_peak, 1),
                                                 round(umb_val, 1)))

        return len(armónicos_detectados) > 0, armónicos_detectados

    except Exception:
        return False, []


# ═══════════════════════════════════════════════════════════════════════════
# MOTOR DE CAPTURA
# ═══════════════════════════════════════════════════════════════════════════

class Motor:
    def __init__(self):
        self.buffer       = collections.deque(maxlen=BUF_N)
        self.lock         = threading.Lock()
        self.stream       = None
        self.activo       = False
        self.modelo_id    = None
        self.modelos_nvh  = {}
        self.umbral_espectro = {}
        self.pinon_manual = None
        self.cb_ciclo     = None
        self.cb_senal     = None
        self.cb_estado    = None
        # ── Historial de identificaciones (filtro de consistencia) ───
        # Guarda las últimas N identificaciones para evitar cambios
        # bruscos de piñón por errores del modelo
        self._hist_pinon  = collections.deque(maxlen=5)
        # ── Carpeta raiz para guardar CSVs ───────────────────────────
        # Carpeta raíz por defecto = Desktop/HORSE (ruta del proyecto)
        _horse = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop", "HORSE")
        self.carpeta_salida = _horse if os.path.isdir(_horse) else os.path.dirname(os.path.abspath(__file__))
        # ── Estado detección de engrane ──────────────────────────────
        self._umbral_rms      = None   # se fija tras calibración real
        self._rms_fondo_real  = None   # valor medido en calibración
        self._cnt_sobre       = 0
        self._cnt_bajo        = 0
        self._hist_rms        = []
        self._engranando      = False
        self._buf_ciclo       = []     # acumula muestras del ciclo (post-engrane)
        self._n_ciclo_obj     = int(DURACION_CICLO * FS)
        # ── Buffer pre-engrane (rolling, siempre activo) ─────────────
        # Almacena los últimos MUESTRAS_PREENGRANE samples en todo momento.
        # Al confirmar engrane se anteponen al ciclo para recuperar el
        # inicio del empuje que ya sonó durante la ventana de confirmación.
        self._buf_pre = collections.deque(maxlen=MUESTRAS_PREENGRANE)

    def cargar_id(self, ruta):
        with open(ruta,"rb") as f: self.modelo_id = pickle.load(f)

    def cargar_nvh(self, ruta):
        with open(ruta,"rb") as f: m = pickle.load(f)
        p = m.get("pinon","?").upper()
        key = "PIMA" if "PIMA" in p else "ARBOL_SEC_15" if "15" in p else "ARBOL_SEC_14"
        self.modelos_nvh[key] = m; return key

    def cargar_umbral_espectro(self, ruta):
        """Carga pkl de umbrales espectrales (calcular_umbrales_espectro.py)."""
        with open(ruta, "rb") as f:
            datos = pickle.load(f)
        # Guardar indexado por piñón para acceso rápido en tiempo real
        pinon = datos.get("pinon", "?").upper()
        key   = "PIMA" if "PIMA" in pinon else "ARBOL_SEC_15" if "15" in pinon else "ARBOL_SEC_14"
        self.umbral_espectro[key] = datos
        return key, datos.get("n_total", "?")

    # ── Calibración real ─────────────────────────────────────────────────────
    def calibrar(self, dev):
        """
        Mide el ruido de fondo usando un InputStream temporal de CALIBRACION_SEG s.
        No usa sd.rec() para evitar conflictos de dispositivo con el stream principal.
        Fija umbral = rms_fondo_real × FACTOR_UMBRAL.
        Si falla, cae al umbral fijo de referencia (UMBRAL_RMS_FIJO).
        Al terminar abre el stream continuo de análisis.
        Debe llamarse desde un hilo secundario (bloquea CALIBRACION_SEG s).
        """
        if self.cb_estado:
            self.cb_estado(
                f"Calibrando ruido de fondo ({int(CALIBRACION_SEG)} s)  —  "
                "no generes ruido cerca del sensor...", "warn")
        try:
            n_cal     = int(CALIBRACION_SEG * FS)
            muestras  = []
            evento    = threading.Event()

            def _cb_cal(indata, frames, time_info, status):
                muestras.extend(indata[:, 0].tolist())
                if len(muestras) >= n_cal:
                    evento.set()

            stream_cal = sd.InputStream(
                device=dev, samplerate=FS, channels=1,
                dtype="float32", blocksize=2048, callback=_cb_cal)
            stream_cal.start()
            evento.wait(timeout=CALIBRACION_SEG + 2.0)
            stream_cal.stop()
            stream_cal.close()

            if len(muestras) < n_cal // 2:
                raise ValueError(f"Pocas muestras grabadas: {len(muestras)}")

            datos     = np.array(muestras[:n_cal], dtype=np.float32)
            rms_fondo = float(np.sqrt(np.mean(datos ** 2)))

            if rms_fondo < 1e-6 or rms_fondo > 1.0:
                raise ValueError(f"RMS de fondo fuera de rango: {rms_fondo:.6f}")

            self._rms_fondo_real = rms_fondo
            self._umbral_rms     = round(rms_fondo * FACTOR_UMBRAL, 6)

            if self.cb_estado:
                self.cb_estado(
                    f"Calibración OK  —  fondo={rms_fondo:.5f}  "
                    f"umbral={self._umbral_rms:.5f} (×{FACTOR_UMBRAL})  "
                    "— escuchando engrane...", "ok")

        except Exception as e:
            self._rms_fondo_real = RMS_FONDO
            self._umbral_rms     = UMBRAL_RMS_FIJO
            if self.cb_estado:
                self.cb_estado(
                    f"Calibración fallida ({e})  —  "
                    f"usando umbral de referencia: {UMBRAL_RMS_FIJO:.5f}", "warn")

        self._iniciar_stream(dev)

    # ── Stream continuo ──────────────────────────────────────────────────────
    def _iniciar_stream(self, dev):
        def cb(indata, frames, time_info, status):
            mono = indata[:,0].astype(np.float32)

            # ── Filtro de spikes del acelerómetro ─────────────────────
            # Reemplaza muestras que superen 8× el RMS local por la
            # mediana del bloque — elimina artefactos de conexión USB
            rms_local = float(np.sqrt(np.mean(mono**2))) + 1e-9
            umbral_spike = rms_local * 8.0
            med = float(np.median(mono))
            mono = np.where(np.abs(mono) > umbral_spike, med, mono)

            with self.lock:
                # Buffer visual (siempre)
                self.buffer.extend(mono.tolist())
                if self._engranando:
                    # Acumular muestras del ciclo activo
                    self._buf_ciclo.extend(mono.tolist())
                else:
                    # Mantener rolling buffer pre-engrane (solo cuando NO engranando)
                    self._buf_pre.extend(mono.tolist())

            if self.cb_senal:
                self.cb_senal(mono)
            self._procesar_bloque(mono)

        for device in ([dev, None] if dev is not None else [None]):
            try:
                self.stream = sd.InputStream(
                    device=device, samplerate=FS, channels=1,
                    dtype="float32", blocksize=2048, callback=cb)
                self.stream.start()
                if self.cb_estado:
                    nombre = (sd.query_devices(device)["name"]
                              if device is not None else "dispositivo por defecto")
                    self.cb_estado(f"Stream activo: {nombre}", "ok")
                return
            except Exception as e:
                if device is None:
                    raise e
                continue

    # ── Detección de engrane bloque a bloque ────────────────────────────────
    def _procesar_bloque(self, mono):
        """
        Detección robusta de engrane con tres filtros en cascada:

        1. RMS sostenido >= umbral durante DURACION_ESTAB (1.20s)
           — descarta golpes cortos del cambio de pieza

        2. CV del historial RMS < MAX_CV durante TODA la ventana
           — descarta señales impulsivas que no son estacionarias

        3. Periodicidad a T_ROT: el RMS por giro debe ser coherente
           (std/mean < 0.5) durante al menos MIN_CICLOS_PERIODO giros
           — confirma que hay rotación real antes de aceptar el engrane
        """
        if self._umbral_rms is None or self._engranando:
            return

        rms = float(np.sqrt(np.mean(mono**2)))
        self._hist_rms.append(rms)
        if len(self._hist_rms) > 50:
            self._hist_rms.pop(0)

        if rms >= self._umbral_rms:
            self._cnt_sobre += len(mono)
            self._cnt_bajo   = 0

            # ── Filtro 2: CV deslizante — exigir estabilidad continua ──
            if len(self._hist_rms) >= 8:
                arr  = np.array(self._hist_rms[-16:])  # ventana deslizante
                mean = np.mean(arr)
                cv   = np.std(arr) / mean if mean > 1e-9 else 999.0
                if cv > MAX_CV:
                    # Señal inestable — resetear y esperar
                    self._cnt_sobre = 0
                    self._cnt_bajo  = 0
                    self._hist_rms.clear()
                    return

            if self._cnt_sobre >= MUESTRAS_ESTAB:
                # ── Filtro 3: Periodicidad a T_ROT ─────────────────────
                # Tomar el buffer pre-engrane y verificar que el RMS
                # por giro sea coherente (engrane real vs golpes)
                with self.lock:
                    buf_check = list(self._buf_pre)

                if len(buf_check) >= M_ROT * MIN_CICLOS_PERIODO:
                    arr_buf = np.array(buf_check, dtype=np.float32)
                    n_giros = len(arr_buf) // M_ROT
                    rms_por_giro = np.array([
                        float(np.sqrt(np.mean(
                            arr_buf[i*M_ROT:(i+1)*M_ROT]**2)))
                        for i in range(n_giros)
                    ])
                    mean_g = np.mean(rms_por_giro)
                    cv_giros = (np.std(rms_por_giro) / mean_g
                                if mean_g > 1e-9 else 999.0)
                    if cv_giros > 0.6:
                        # RMS por giro muy variable — no es rotación estable
                        self._cnt_sobre = 0
                        self._cnt_bajo  = 0
                        self._hist_rms.clear()
                        if self.cb_estado:
                            self.cb_estado(
                                f"Golpe descartado — CV_giros={cv_giros:.2f} "
                                f"(umbral 0.6) — no es engrane", "warn")
                        return

                # ── Engrane confirmado (pasó los 3 filtros) ────────────
                self._cnt_sobre = 0
                self._cnt_bajo  = 0
                self._hist_rms.clear()

                with self.lock:
                    pre = list(self._buf_pre)
                    self._buf_pre.clear()
                    self._buf_ciclo = pre

                self._engranando = True
                if self.cb_estado:
                    n_pre_s = len(pre) / FS
                    self.cb_estado(
                        f"Engrane confirmado — pre={n_pre_s:.2f}s  "
                        "acumulando ciclo...", "warn")
                threading.Thread(target=self._esperar_ciclo, daemon=True).start()
        else:
            self._cnt_bajo += len(mono)
            if self._cnt_bajo >= MUESTRAS_SILENCIO:
                self._cnt_sobre = 0
                self._hist_rms.clear()

    # ── Esperar ciclo completo y procesar ────────────────────────────────────
    def _esperar_ciclo(self):
        """
        Espera a que _buf_ciclo acumule DURACION_CICLO s de señal.
        El buffer ya contiene las muestras pre-engrane al inicio,
        por lo que el ciclo entregado a analizar_ciclo incluye el
        inicio del empuje desde antes de la confirmación.
        Timeout de seguridad: 10 s.
        """
        try:
            t_inicio = time.time()
            while True:
                with self.lock:
                    n_actual = len(self._buf_ciclo)
                if n_actual >= self._n_ciclo_obj:
                    break
                if time.time() - t_inicio > 10.0:
                    if self.cb_estado:
                        self.cb_estado("Timeout esperando ciclo completo", "warn")
                    return
                time.sleep(0.05)

            with self.lock:
                senal = np.array(self._buf_ciclo[:self._n_ciclo_obj],
                                 dtype=np.float32)

            # ── Identificar piñón ──────────────────────────────────
            # Identificación siempre manual — el operador selecciona el piñón
            # presionando el cuadro en la ventana de semáforo.
            if self.pinon_manual:
                pk, ci = self.pinon_manual, 1.0
            else:
                # Sin selección manual → no procesar, avisar al operador
                if self.cb_estado:
                    self.cb_estado(
                        "Selecciona el piñón en la pantalla de diagnóstico "
                        "antes de iniciar.", "warn")
                return
            if False:  # bloque inactivo — identificación automática deshabilitada
                pk, ci = identificar_pinon(senal, self.modelo_id,
                                           self._hist_pinon)

            # ── Segmentar y calcular features ──────────────────────
            feats = analizar_ciclo(senal)
            if feats is None:
                if self.cb_estado:
                    self.cb_estado("Ciclo no válido — segmentación fallida", "warn")
                return

            # ── Clasificar NVH ─────────────────────────────────────
            etq, conf = clasificar_nvh(feats, pk, self.modelos_nvh)

            # ── Guardar CSV con señal cruda ────────────────────────
            # Estructura: DIV-33_DD-MM-YYYY / PIMA_DD-MM-YYYY o AS14_DD-MM-YYYY
            fecha_hoy = datetime.datetime.now().strftime("%d-%m-%Y")
            carpeta_dia = f"DIV-33_{fecha_hoy}"

            if "PIMA" in pk.upper():
                carpeta_pinon = f"PIMA_{fecha_hoy}"
            elif "15" in pk:
                carpeta_pinon = f"AS15_{fecha_hoy}"
            else:
                carpeta_pinon = f"AS14_{fecha_hoy}"

            ruta_carpeta = os.path.join(self.carpeta_salida, carpeta_dia, carpeta_pinon)
            try:
                os.makedirs(ruta_carpeta, exist_ok=True)
            except Exception as e_dir:
                fallback = os.path.join(os.path.expanduser("~"),
                                        "OneDrive", "Desktop", "DATOS_MACHINE_LEARNING",
                                        carpeta_dia, carpeta_pinon)
                os.makedirs(fallback, exist_ok=True)
                ruta_carpeta = fallback
                if self.cb_estado:
                    self.cb_estado(
                        f"Carpeta inaccesible ({e_dir}) — "
                        f"guardando en: {fallback}", "warn")

            ts       = datetime.datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:20]
            nombre   = f"engrane_{ts}_{etq}.csv"
            ruta_csv = os.path.join(ruta_carpeta, nombre)
            t_arr    = np.arange(len(senal)) / FS
            import pandas as _pd
            _pd.DataFrame({"tiempo_s": t_arr, "senal": senal}).to_csv(ruta_csv, index=False)

            res = {
                "hora":      datetime.datetime.now().strftime("%H:%M:%S"),
                "fecha":     datetime.datetime.now().strftime("%d/%m/%Y"),
                "pinon":     pk,     "conf_id":   round(ci,   2),
                "etiqueta":  etq,    "confianza": round(conf, 2),
                "K_ret":     feats["K_ret"],      "CF_ret":  feats["CF_p99_ret"],
                "K_emp":     feats["K_emp"],       "CF_emp":  feats["CF_p99_emp"],
                "rms_ret":   feats["rms_ret"],     "rms_emp": feats["rms_emp"],
                "dur_ret":   feats["dur_ret"],
                "csv":       ruta_csv,
                "seg_ret":   feats.get("seg_ret"),
                "seg_emp":   feats.get("seg_emp"),
            }
            # ── Detección GOLPE y SIRENEO ──────────────────────────────
            golpe, zona_golpe       = detectar_golpe(feats, pk, self.modelos_nvh)
            sireneo, armon_sireneo  = detectar_sireneo(
                feats, pk, self.umbral_espectro)
            # Avisar si no hay pkl cargado para este piñón
            if not self.umbral_espectro.get(pk):
                if self.cb_estado:
                    self.cb_estado(
                        f"⚠ Sin umbrales espectrales para {pk} — "
                        f"sireneo no disponible. Carga el .pkl.", "warn")
            res["golpe"]            = golpe
            res["zona_golpe"]       = zona_golpe
            res["sireneo"]          = sireneo
            res["armon_sireneo"]    = armon_sireneo
            if self.cb_ciclo:
                self.cb_ciclo(res)
            if self.cb_estado:
                self.cb_estado(
                    f"Guardado: {nombre}  —  {etq}  "
                    f"K={feats['K_ret']}  CF={feats['CF_p99_ret']}",
                    "ok" if etq == "BUENO" else "err")

        except Exception as e:
            import traceback
            traceback.print_exc()
            if self.cb_estado:
                self.cb_estado(f"Error en ciclo: {e}", "err")
        finally:
            self._engranando = False

    # ── Arranque y parada ────────────────────────────────────────────────────
    def iniciar(self, dev):
        """
        Lanza calibración real en hilo separado para no bloquear la UI.
        El stream arranca automáticamente al terminar la calibración.
        """
        if self.activo:
            return
        self.activo = True
        threading.Thread(
            target=self.calibrar, args=(dev,), daemon=True).start()

    def detener(self):
        self.activo      = False
        self._engranando = False
        self._umbral_rms = None
        self._cnt_sobre  = 0
        self._cnt_bajo   = 0
        self._buf_pre.clear()
        self._hist_pinon.clear()
        if self.stream:
            try:
                self.stream.stop()
                self.stream.close()
            except Exception:
                pass
            self.stream = None

# ═══════════════════════════════════════════════════════════════════════════
# EXPORTAR EXCEL
# ═══════════════════════════════════════════════════════════════════════════

def exportar_excel(ciclos, operador, turno, ruta):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Turno"
    t  = Side(style="thin", color="CBD5E1"); bord = Border(left=t,right=t,top=t,bottom=t)
    fh = PatternFill("solid", fgColor="1E293B")
    fm = PatternFill("solid", fgColor="FEE2E2")
    fr = PatternFill("solid", fgColor="FEF9C3")
    fb = PatternFill("solid", fgColor="DCFCE7")
    fa = PatternFill("solid", fgColor="F8FAFC")

    ws.merge_cells("A1:N1")
    ws["A1"] = (f"HISTORIAL DE TURNO  —  {operador.upper()}  |  "
                f"Turno: {turno}  |  "
                f"{datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}")
    ws["A1"].font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    ws["A1"].fill = fh; ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    HDRS   = ["N°","Fecha","Hora","Piñón","Conf ID","Resultado","Conf ML",
              "K ret","CF ret","K emp","CF emp","RMS ret","RMS emp","Dur (s)"]
    WIDTHS = [5,12,10,14,8,12,8,8,8,8,8,10,10,10]
    for ci,(h,w) in enumerate(zip(HDRS,WIDTHS),1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = Font(name="Arial",bold=True,size=9,color="FFFFFF")
        c.fill = fh; c.alignment = Alignment(horizontal="center"); c.border = bord
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
    ws.row_dimensions[2].height = 18

    for ri, cc in enumerate(ciclos, 3):
        # Determinar resultado basado en golpe/sireneo
        golpe   = cc.get("golpe",   False)
        sireneo = cc.get("sireneo", False)
        if sireneo:
            resultado = "DESECHAR"
            fill      = fm
            etq_col   = "991B1B"
        elif golpe:
            resultado = "REPARAR"
            fill      = fr
            etq_col   = "854D0E"
        else:
            resultado = "BUENO"
            fill      = fb
            etq_col   = "166534"

        vals = [ri-2,
                cc.get("fecha", ""),
                cc.get("hora",  ""),
                cc.get("pinon", ""),
                cc.get("conf_id", ""),
                resultado,
                cc.get("confianza", ""),
                cc.get("K_ret",   ""),
                cc.get("CF_ret",  ""),
                cc.get("K_emp",   ""),
                cc.get("CF_emp",  ""),
                cc.get("rms_ret", ""),
                cc.get("rms_emp", ""),
                cc.get("dur_ret", "")]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = Font(name="Arial", size=8, bold=(ci==6),
                          color=etq_col if ci==6 else "000000")
            c.fill = fill; c.border = bord
            c.alignment = Alignment(horizontal="center")
        ws.row_dimensions[ri].height = 13

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:N{len(ciclos)+2}"

    # Hoja Resumen
    ws2 = wb.create_sheet("Resumen")
    ws2.column_dimensions["A"].width = 20; ws2.column_dimensions["B"].width = 14
    ws2.merge_cells("A1:B1"); ws2["A1"] = "RESUMEN DEL TURNO"
    ws2["A1"].font = Font(name="Arial",bold=True,size=12,color="FFFFFF")
    ws2["A1"].fill = fh
    ws2["A1"].alignment = Alignment(horizontal="center")

    n_bueno   = sum(1 for c in ciclos
                    if not c.get("golpe") and not c.get("sireneo"))
    n_reparar = sum(1 for c in ciclos
                    if c.get("golpe") and not c.get("sireneo"))
    n_desechar = sum(1 for c in ciclos if c.get("sireneo"))

    for ri,(k,v) in enumerate([
        ("Operador",  operador),
        ("Turno",     turno),
        ("Total",     len(ciclos)),
        ("BUENO",     n_bueno),
        ("REPARAR",   n_reparar),
        ("DESECHAR",  n_desechar),
        ("Inicio",    ciclos[0].get("hora",  "—") if ciclos else "—"),
        ("Fin",       ciclos[-1].get("hora", "—") if ciclos else "—")], 2):
        ws2.cell(row=ri, column=1, value=k).font = Font(name="Arial", bold=True, size=9)
        ws2.cell(row=ri, column=2, value=v).font = Font(name="Arial", size=9)
        for ci in [1,2]:
            ws2.cell(row=ri, column=ci).border = bord

    wb.save(ruta)

# ═══════════════════════════════════════════════════════════════════════════
# HELPER — boton estilo explorador
# ═══════════════════════════════════════════════════════════════════════════

def hacer_boton(parent, texto, cmd, bg=C_SURFACE2, fg=C_TEXT, ancho=18, bold=False):
    return tk.Button(parent, text=texto, command=cmd, bg=bg, fg=fg,
                     activebackground=bg, activeforeground=fg,
                     relief="flat", bd=0,
                     font=(C_MONO, 9, "bold" if bold else "normal"),
                     width=ancho, cursor="hand2")

# ═══════════════════════════════════════════════════════════════════════════
# VENTANA 2 — TABLA DE TURNO
# ═══════════════════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════════════════
# VENTANA SEMÁFORO — círculo BUENO / MALO estilo botón emergencia
# ═══════════════════════════════════════════════════════════════════════════

class VentanaSemaforo(tk.Toplevel):

    def __init__(self, master):
        super().__init__(master)
        self.title("MAT IA — Estado")
        self.geometry("520x480")
        self.minsize(420, 400)
        self.configure(bg=C_BG)
        _aplicar_icono(self)
        self._golpe      = False
        self._sireneo    = False
        self._zona_golpe = ""
        self._n_golpe    = 0
        self._n_sireneo  = 0
        self._n_bueno    = 0
        self._timer_reset = None
        self._build()

    def _build(self):
        # ── Header ──────────────────────────────────────────────────────
        hdr = tk.Frame(self, bg=C_SURFACE, height=56)
        hdr.pack(fill="x"); hdr.pack_propagate(False)
        tk.Frame(hdr, bg=C_ACENTO, width=4).pack(side="left", fill="y")
        _logo = _cargar_logo()
        if _logo:
            tk.Label(hdr, image=_logo, bg=C_SURFACE).pack(side="left", padx=14, pady=8)
        else:
            tk.Label(hdr, text="HORSE", bg=C_SURFACE, fg=C_ACENTO,
                     font=(C_MONO, 13, "bold")).pack(side="left", padx=14)
        tk.Frame(hdr, bg=C_BORDER, width=1).pack(side="left", fill="y", pady=8)
        fh = tk.Frame(hdr, bg=C_SURFACE); fh.pack(side="left", padx=12)
        tk.Label(fh, text="DIAGNÓSTICO DEL PIÑÓN", bg=C_SURFACE,
                 fg=C_TEXT, font=(C_MONO, 11, "bold")).pack(anchor="w")
        self.lbl_sub = tk.Label(fh, text="En espera...",
                                 bg=C_SURFACE, fg=C_TEXT_SUB, font=(C_MONO, 9))
        self.lbl_sub.pack(anchor="w")
        _lm = _hacer_logo_matia(hdr, bg=C_SURFACE)
        _lm.pack(side="right", padx=14, pady=6)
        tk.Frame(self, bg=C_BORDER, height=1).pack(fill="x")

        # ── Zona superior: selector izquierda + cuadro central ──────────
        C_BTN_ON  = "#1a5fa8"
        C_BTN_OFF = C_SURFACE2
        C_TXT_ON  = "white"
        C_TXT_OFF = C_TEXT_SUB

        self._var_manual = tk.StringVar(value="")
        self._cb_manual  = None

        img_pima  = _cargar_img_pinon("PIMA",  ancho=180, alto=140)
        img_arbol = _cargar_img_pinon("ARBOL", ancho=260, alto=130)

        def _sel_pinon(key):
            self._var_manual.set(key)
            if hasattr(self, "_cb_manual") and self._cb_manual:
                self._cb_manual(key)
            _actualizar_selector()

        def _actualizar_selector():
            sel = self._var_manual.get()
            btn_pima.config(
                bg=C_BTN_ON if sel == "PIMA" else C_BTN_OFF,
                fg=C_TXT_ON if sel == "PIMA" else C_TXT_OFF)
            btn_as.config(
                bg=C_BTN_ON if sel == "ARBOL_SEC_14" else C_BTN_OFF,
                fg=C_TXT_ON if sel == "ARBOL_SEC_14" else C_TXT_OFF)
            if sel == "PIMA":
                if img_pima:
                    self._lbl_img_central.config(image=img_pima, text="")
                    self._lbl_img_central.image = img_pima
                else:
                    self._lbl_img_central.config(image="", text="PIMA")
                self._lbl_nombre_central.config(text="PIMA  ·  26 dientes")
                self._lbl_estado_central.config(text="▶ EN PRODUCCIÓN", fg=C_BTN_ON)
            elif sel == "ARBOL_SEC_14":
                if img_arbol:
                    self._lbl_img_central.config(image=img_arbol, text="")
                    self._lbl_img_central.image = img_arbol
                else:
                    self._lbl_img_central.config(image="", text="ÁRBOL SEC.")
                self._lbl_nombre_central.config(text="Árbol Secundario  ·  14 dientes")
                self._lbl_estado_central.config(text="▶ EN PRODUCCIÓN", fg=C_BTN_ON)
            else:
                self._lbl_img_central.config(image="", text="Selecciona\nun piñón")
                self._lbl_img_central.image = None
                self._lbl_nombre_central.config(text="")
                self._lbl_estado_central.config(text="Sin selección", fg=C_TEXT_DIM)

        # Contenedor: col0=botones  col1=sep  col2=cuadro central
        f_zona = tk.Frame(self, bg=C_BG)
        f_zona.pack(fill="x", padx=16, pady=(14, 8))
        f_zona.columnconfigure(0, weight=0)
        f_zona.columnconfigure(1, weight=0)
        f_zona.columnconfigure(2, weight=1)

        # ── Botones izquierda ─────────────────────────────────────────────
        f_btns = tk.Frame(f_zona, bg=C_BG)
        f_btns.grid(row=0, column=0, sticky="ns", padx=(0, 0))
        tk.Label(f_btns, text="PIÑÓN", bg=C_BG, fg=C_TEXT_SUB,
                 font=("Arial", 8, "bold")).pack(anchor="w", pady=(4, 6))

        btn_pima = tk.Button(f_btns,
            text="PIMA" + chr(10) + "26 dientes",
            bg=C_BTN_OFF, fg=C_TXT_OFF,
            activebackground=C_BTN_ON, activeforeground="white",
            relief="flat", bd=0, cursor="hand2",
            font=("Arial", 10, "bold"), width=14, height=3,
            command=lambda: _sel_pinon("PIMA"))
        btn_pima.pack(fill="x", pady=(0, 8))

        btn_as = tk.Button(f_btns,
            text="ÁRBOL" + chr(10) + "SECUNDARIO" + chr(10) + "14 dientes",
            bg=C_BTN_OFF, fg=C_TXT_OFF,
            activebackground=C_BTN_ON, activeforeground="white",
            relief="flat", bd=0, cursor="hand2",
            font=("Arial", 10, "bold"), width=14, height=4,
            command=lambda: _sel_pinon("ARBOL_SEC_14"))
        btn_as.pack(fill="x")

        # ── Separador vertical ────────────────────────────────────────────
        tk.Frame(f_zona, bg=C_BORDER, width=1
                 ).grid(row=0, column=1, sticky="ns", padx=12, pady=4)

        # ── Cuadro central único ──────────────────────────────────────────
        self.f_cuadro_central = tk.Frame(f_zona, bg=C_SURFACE2,
            highlightbackground=C_BORDER, highlightthickness=2, relief="flat")
        self.f_cuadro_central.grid(row=0, column=2, sticky="nsew", ipady=10)

        self._lbl_img_central = tk.Label(self.f_cuadro_central,
            text="Selecciona\nun piñón", bg=C_SURFACE2, fg=C_TEXT_DIM,
            font=("Arial", 12))
        self._lbl_img_central.pack(expand=True, pady=(12, 4))

        self._lbl_nombre_central = tk.Label(self.f_cuadro_central, text="",
            bg=C_SURFACE2, fg=C_TEXT, font=("Arial", 10, "bold"))
        self._lbl_nombre_central.pack()

        self._lbl_estado_central = tk.Label(self.f_cuadro_central,
            text="Sin selección", bg=C_SURFACE2, fg=C_TEXT_DIM,
            font=("Arial", 8))
        self._lbl_estado_central.pack(pady=(2, 10))

        # Alias para compatibilidad con métodos existentes
        self.f_cuadro_pima       = self.f_cuadro_central
        self.f_cuadro_as         = self.f_cuadro_central
        self.lbl_pima_estado     = self._lbl_estado_central
        self.lbl_as_estado       = self._lbl_estado_central
        self._lbl_img_pima_ref   = None
        self._lbl_img_arbol_ref  = None

        # ── Dos círculos lado a lado ─────────────────────────────────────
        f_circles = tk.Frame(self, bg=C_BG)
        f_circles.pack(fill="both", expand=True, padx=16, pady=(4, 8))

        # Círculo GOLPE (izquierda)
        f_g = tk.Frame(f_circles, bg=C_BG)
        f_g.pack(side="left", fill="both", expand=True)
        self.canvas_golpe = tk.Canvas(f_g, bg=C_BG, highlightthickness=0)
        self.canvas_golpe.pack(fill="both", expand=True)
        self.lbl_golpe_cnt = tk.Label(f_g, text="GOLPE  0",
                                       bg=C_BG, fg=C_TEXT_DIM,
                                       font=(C_MONO, 9, "bold"))
        self.lbl_golpe_cnt.pack(pady=(4, 0))
        self.canvas_golpe.bind("<Configure>", lambda e: self._dibujar_circulo(
            self.canvas_golpe, "GOLPE", self._golpe))

        # Separador
        tk.Frame(f_circles, bg=C_BORDER2, width=1).pack(side="left", fill="y", pady=20)

        # Círculo SIRENEO (derecha)
        f_s = tk.Frame(f_circles, bg=C_BG)
        f_s.pack(side="left", fill="both", expand=True)
        self.canvas_sireneo = tk.Canvas(f_s, bg=C_BG, highlightthickness=0)
        self.canvas_sireneo.pack(fill="both", expand=True)
        self.lbl_sireneo_cnt = tk.Label(f_s, text="SIRENEO  0",
                                         bg=C_BG, fg=C_TEXT_DIM,
                                         font=(C_MONO, 9, "bold"))
        self.lbl_sireneo_cnt.pack(pady=(4, 0))
        self.canvas_sireneo.bind("<Configure>", lambda e: self._dibujar_circulo(
            self.canvas_sireneo, "SIRENEO", self._sireneo))

        # ── Barra inferior ───────────────────────────────────────────────
        tk.Frame(self, bg=C_BORDER, height=1).pack(fill="x")
        f_st = tk.Frame(self, bg=C_SURFACE, height=30)
        f_st.pack(fill="x"); f_st.pack_propagate(False)
        self.lbl_pinon = tk.Label(f_st, text="Piñón: —", bg=C_SURFACE,
                                   fg=C_TEXT_DIM, font=(C_MONO, 8), anchor="w")
        self.lbl_pinon.pack(side="left", padx=12)
        self.lbl_detalle = tk.Label(f_st, text="", bg=C_SURFACE,
                                     fg=C_TEXT_DIM, font=(C_MONO, 8), anchor="w")
        self.lbl_detalle.pack(side="left", padx=8)

    def _dibujar_circulo(self, cv, tipo, estado):
        """
        estado="activo"  → alerta (rojo para GOLPE, ámbar para SIRENEO)
        estado="ok"      → verde — ciclo analizado sin ese defecto
        estado="espera"  → gris  — sin ciclo activo
        También acepta bool True/False por compatibilidad.
        """
        # Compatibilidad con llamadas que pasan bool
        if estado is True:  estado = "activo"
        if estado is False: estado = "espera"

        cv.delete("all")
        w = cv.winfo_width(); h = cv.winfo_height()
        if w < 10 or h < 10: return
        cx = w // 2; cy = h // 2
        r  = min(w, h) // 2 - 20

        if estado == "activo":
            if tipo == "GOLPE":
                c_fill = "#991b1b"; c_glow = "#2b0d0d"; c_bord = C_MALO
            else:
                c_fill = "#991b1b"; c_glow = "#2b0d0d"; c_bord = C_MALO
            cv.create_oval(cx-r-16, cy-r-16, cx+r+16, cy+r+16,
                           fill=c_glow, outline="")
            cv.create_oval(cx-r-6, cy-r-6, cx+r+6, cy+r+6,
                           fill=c_bord, outline="")
            cv.create_oval(cx-r, cy-r, cx+r, cy+r,
                           fill=c_fill, outline="")

            if tipo == "GOLPE":
                # Determinar texto de giro según flanco
                zona = getattr(self, "_zona_golpe", "")
                if zona == "EMP":
                    txt_giro = "PRIMER GIRO"
                elif zona == "RET":
                    txt_giro = "SEGUNDO GIRO"
                elif zona in ("RET+EMP", "EMP+RET"):
                    txt_giro = "AMBOS GIROS"
                else:
                    txt_giro = "REPARAR"

                cv.create_text(cx, cy - r//4, text="GOLPE EN",
                               fill="#ffffff",
                               font=(C_MONO, max(10, r//4), "bold"))
                cv.create_text(cx, cy + r//6, text=txt_giro,
                               fill="#fca5a5",
                               font=(C_MONO, max(8, r//5), "bold"))
            else:
                # SIRENEO
                cv.create_text(cx, cy - r//6, text=tipo,
                               fill="#ffffff",
                               font=(C_MONO, max(10, r//4), "bold"))
                cv.create_text(cx, cy + r//4, text="DESECHAR",
                               fill="#ffffff",
                               font=(C_MONO, max(9, r//5), "normal"))

        elif estado == "ok":
            cv.create_oval(cx-r-16, cy-r-16, cx+r+16, cy+r+16,
                           fill="#0d2b18", outline="")
            cv.create_oval(cx-r-6, cy-r-6, cx+r+6, cy+r+6,
                           fill=C_BUENO, outline="")
            cv.create_oval(cx-r, cy-r, cx+r, cy+r,
                           fill="#15803d", outline="")
            cv.create_text(cx, cy, text="OK",
                           fill="#ffffff",
                           font=(C_MONO, max(13, r//3), "bold"))

        else:  # espera
            cv.create_oval(cx-r-6, cy-r-6, cx+r+6, cy+r+6,
                           fill="#9ca3af", outline="")
            cv.create_oval(cx-r, cy-r, cx+r, cy+r,
                           fill="#d1d5db", outline="")
            cv.create_text(cx, cy - r//6, text=tipo,
                           fill="#6b7280",
                           font=(C_MONO, max(11, r//4), "bold"))
            cv.create_text(cx, cy + r//4, text="EN ESPERA",
                           fill="#9ca3af",
                           font=(C_MONO, max(7, r//7), "normal"))

    def actualizar(self, golpe, sireneo, pinon=None, zona_golpe="",
                   armon_sireneo=None):
        self._golpe      = golpe
        self._sireneo    = sireneo
        self._zona_golpe = zona_golpe   # guardar para _dibujar_circulo

        if golpe:
            self._n_golpe += 1
        if sireneo:
            self._n_sireneo += 1
        if not golpe and not sireneo:
            self._n_bueno += 1

        # Contadores — rojo si activo, verde si OK
        self.lbl_golpe_cnt.config(
            text=f"GOLPE  {self._n_golpe}",
            fg=C_MALO if golpe else C_BUENO)
        self.lbl_sireneo_cnt.config(
            text=f"SIRENEO  {self._n_sireneo}",
            fg=C_MALO if sireneo else C_BUENO)

        # Subtítulo
        partes = []
        if golpe:
            # Traducir zona a terminología de operador
            if zona_golpe == "EMP":
                txt_zona = "GOLPE EN PRIMER GIRO"
            elif zona_golpe == "RET":
                txt_zona = "GOLPE EN SEGUNDO GIRO"
            elif zona_golpe in ("RET+EMP", "EMP+RET"):
                txt_zona = "GOLPE EN AMBOS GIROS"
            else:
                txt_zona = "GOLPE DETECTADO"
            partes.append(txt_zona)
        if sireneo and armon_sireneo:
            ks = "+".join(f"{k}×GMF({f})" for k, f, _, _ in armon_sireneo[:2])
            partes.append(f"SIRENEO {ks}")
        if not partes: partes.append("OK — sin defectos")
        self.lbl_sub.config(
            text=" | ".join(partes),
            fg=C_MALO if (golpe or sireneo) else C_BUENO)

        if pinon:
            self.lbl_pinon.config(text=f"Piñón: {pinon}", fg=C_ACENTO)
            # Iluminar cuadro del piñón activo en verde
            es_pima = "PIMA" in pinon.upper()
            # PIMA
            bg_p  = "#F59E0B" if es_pima  else C_SURFACE2
            brd_p = "#D97706" if es_pima  else C_BORDER
            txt_p = "▶ EN PROCESO" if es_pima else ""
            # AS14
            bg_a  = "#F59E0B" if not es_pima else C_SURFACE2
            brd_a = "#D97706" if not es_pima else C_BORDER
            txt_a = "▶ EN PROCESO" if not es_pima else ""
            for w in self.f_cuadro_pima.winfo_children():
                w.config(bg=bg_p)
            self.f_cuadro_pima.config(bg=bg_p,
                                       highlightbackground=brd_p)
            self.lbl_pima_estado.config(text=txt_p, bg=bg_p,
                                         fg="white" if es_pima else C_BUENO)
            for w in self.f_cuadro_as.winfo_children():
                w.config(bg=bg_a)
            self.f_cuadro_as.config(bg=bg_a,
                                     highlightbackground=brd_a)
            self.lbl_as_estado.config(text=txt_a, bg=bg_a,
                                       fg="white" if not es_pima else C_BUENO)

        detalle = ""
        if sireneo and armon_sireneo:
            detalle = "  ".join(
                f"×{k}{fl[0].upper()}={v:.0f}dB(umb={u:.0f})"
                for k, fl, v, u in armon_sireneo)
        self.lbl_detalle.config(text=detalle,
                                 fg=C_MALO if sireneo else C_TEXT_DIM)

        # Dibujar círculos: rojo si activo, verde si no
        # Dibujar círculos con tres estados
        self._dibujar_circulo(self.canvas_golpe,   "GOLPE",
                              "activo" if golpe   else "ok")
        self._dibujar_circulo(self.canvas_sireneo, "SIRENEO",
                              "activo" if sireneo else "ok")

        # El resultado se mantiene visible hasta el próximo engrane.
        # _volver_espera() se llama desde VentanaOperador al detectar engrane.

    def _volver_espera(self):
        """Vuelve ambos círculos a EN ESPERA (gris)."""
        self._timer_reset = None
        self._golpe      = False
        self._sireneo    = False
        self._zona_golpe = ""
        self.lbl_sub.config(text="En espera...", fg=C_TEXT_DIM)
        self.lbl_golpe_cnt.config(fg=C_TEXT_DIM)
        self.lbl_sireneo_cnt.config(fg=C_TEXT_DIM)
        self.lbl_detalle.config(text="")
        self._dibujar_circulo(self.canvas_golpe,   "GOLPE",   "espera")
        self._dibujar_circulo(self.canvas_sireneo, "SIRENEO", "espera")

    def iniciar_espera(self):
        """Llamado cuando se detecta nuevo engrane — pone círculos en EN ESPERA
        mientras se procesa el ciclo. El cuadro del piñón mantiene su color.
        """
        if not self.winfo_exists():
            return
        self._golpe      = False
        self._sireneo    = False
        self._zona_golpe = ""
        self.lbl_sub.config(text="Engranando — analizando...", fg=C_REVISAR)
        self.lbl_detalle.config(text="")
        self._dibujar_circulo(self.canvas_golpe,   "GOLPE",   "espera")
        self._dibujar_circulo(self.canvas_sireneo, "SIRENEO", "espera")


class VentanaTurno(tk.Toplevel):
    COLS   = ["N°","Hora","Piñón","Resultado","K ret","CF ret",
              "K emp","CF emp","RMS ret","RMS emp","Dur (s)"]
    WIDTHS = [42, 80, 124, 92, 72, 72, 72, 72, 86, 86, 66]

    def __init__(self, master, operador, turno):
        super().__init__(master)
        self.title("MAT IA — Tabla de Turno")
        self.geometry("1060x660"); self.minsize(800,400)
        self.configure(bg=C_BG)
        _aplicar_icono(self)
        self.operador = operador; self.turno = turno
        self.ciclos = []; self._n = 0
        # Persistencia — JSON junto al ejecutable
        _base = (os.path.dirname(sys.executable)
                 if getattr(sys, "frozen", False)
                 else os.path.dirname(os.path.abspath(__file__)))
        self._ruta_ciclos = os.path.join(_base, "ciclos_turno.json")
        self._build()
        self._cargar_ciclos_persistidos()

    def _build(self):
        # ── Header idéntico al explorador ────────────────────────────────
        frame_hdr = tk.Frame(self, bg=C_SURFACE, height=56)
        frame_hdr.pack(fill="x"); frame_hdr.pack_propagate(False)
        tk.Frame(frame_hdr, bg=C_ACENTO, width=4).pack(side="left", fill="y")
        _logo = _cargar_logo()
        if _logo:
            tk.Label(frame_hdr, image=_logo, bg=C_SURFACE
                     ).pack(side="left", padx=14, pady=8)
        else:
            tk.Label(frame_hdr, text="HORSE", bg=C_SURFACE, fg=C_ACENTO,
                     font=(C_MONO, 13, "bold")).pack(side="left", padx=14)
        tk.Frame(frame_hdr, bg=C_BORDER, width=1).pack(side="left", fill="y", pady=8)
        fh = tk.Frame(frame_hdr, bg=C_SURFACE); fh.pack(side="left", padx=12)
        tk.Label(fh, text="TABLA DE TURNO  —  DEMM", bg=C_SURFACE,
                 fg=C_TEXT, font=(C_MONO,11,"bold")).pack(anchor="w")
        self.lbl_info_hdr = tk.Label(fh,
            text=f"Operador: {self.operador}  ·  Turno: {self.turno}  ·  0 ciclos",
            bg=C_SURFACE, fg=C_TEXT_SUB, font=(C_MONO,9))
        self.lbl_info_hdr.pack(anchor="w")
        # Contadores en header derecho
        fh2 = tk.Frame(frame_hdr, bg=C_SURFACE); fh2.pack(side="right", padx=18)
        self.lbl_nb = tk.Label(fh2, text="BUENO   0", bg=C_SURFACE,
                                fg=C_BUENO, font=(C_MONO,10,"bold"))
        self.lbl_nb.pack(anchor="e")
        self.lbl_nm = tk.Label(fh2, text="MALO    0", bg=C_SURFACE,
                                fg=C_MALO,  font=(C_MONO,10,"bold"))
        self.lbl_nm.pack(anchor="e")
        self.lbl_nr = tk.Label(fh2, text="REVISAR 0", bg=C_SURFACE,
                                fg=C_REVISAR, font=(C_MONO,10,"bold"))
        self.lbl_nr.pack(anchor="e")

        # ── Logo MatIA — esquina superior derecha ─────────────────
        _lm = _hacer_logo_matia(frame_hdr, bg=C_SURFACE)
        _lm.pack(side="right", padx=14, pady=6)
        tk.Frame(self, bg=C_BORDER, height=1).pack(fill="x")

        # ── Layout: lista izquierda + tabla derecha ────────────────────
        frame_main = tk.Frame(self, bg=C_BG); frame_main.pack(fill="both", expand=True)

        # Panel izquierdo — resumen turno
        frame_izq = tk.Frame(frame_main, bg=C_SURFACE, width=200)
        frame_izq.pack(side="left", fill="y"); frame_izq.pack_propagate(False)
        tk.Frame(frame_main, bg=C_BORDER, width=1).pack(side="left", fill="y")

        tk.Label(frame_izq, text="TURNO", bg=C_SURFACE, fg=C_TEXT_SUB,
                 font=(C_MONO,8,"bold"), padx=12, pady=8).pack(anchor="w")
        tk.Frame(frame_izq, bg=C_BORDER, height=1).pack(fill="x")

        self.lbl_op_izq  = tk.Label(frame_izq, text=f"  Op: {self.operador}",
                                     bg=C_SURFACE, fg=C_TEXT_SUB, font=(C_MONO,8),
                                     anchor="w"); self.lbl_op_izq.pack(fill="x", pady=2)
        self.lbl_tur_izq = tk.Label(frame_izq, text=f"  Turno: {self.turno}",
                                     bg=C_SURFACE, fg=C_TEXT_SUB, font=(C_MONO,8),
                                     anchor="w"); self.lbl_tur_izq.pack(fill="x")
        tk.Frame(frame_izq, bg=C_BORDER, height=1).pack(fill="x", pady=(8,0))

        tk.Label(frame_izq, text="ÚLTIMO CICLO", bg=C_SURFACE, fg=C_TEXT_SUB,
                 font=(C_MONO,8,"bold"), padx=12, pady=6).pack(anchor="w")
        tk.Frame(frame_izq, bg=C_BORDER, height=1).pack(fill="x")

        # Recuadros diagnóstico — estilo idéntico al explorador
        def _hacer_recuadro(titulo):
            f = tk.Frame(frame_izq, bg="#f8f9fb", relief="flat", bd=0,
                         padx=12, pady=6,
                         highlightbackground=C_BORDER2, highlightthickness=2)
            f.pack(fill="x", padx=8, pady=4)
            tk.Label(f, text=titulo, bg="#f8f9fb", fg=C_TEXT_SUB,
                     font=(C_MONO,7,"bold")).grid(row=0, column=0, columnspan=3,
                                                   sticky="w", pady=(0,3))
            def _fila(row, etiq):
                tk.Label(f, text=etiq, bg="#f8f9fb", fg=C_TEXT_SUB,
                         font=(C_MONO,8)).grid(row=row, column=0, sticky="e", padx=(0,3))
                lv = tk.Label(f, text="—", bg="#f8f9fb", fg=C_TEXT,
                               font=(C_MONO,10,"bold"), width=6, anchor="w")
                lv.grid(row=row, column=1, sticky="w")
                le = tk.Label(f, text="", bg="#f8f9fb", fg=C_TEXT_SUB,
                               font=(C_MONO,8), width=8, anchor="w")
                le.grid(row=row, column=2, sticky="w")
                return lv, le
            lk, lke = _fila(1, "K =")
            lc, lce = _fila(2, "CF.p99 =")
            return f, lk, lke, lc, lce

        (self.f_ret, self.lbl_k_ret, self.lbl_k_ret_e,
                     self.lbl_cf_ret, self.lbl_cf_ret_e) = _hacer_recuadro("── RETROCESO ──")
        tk.Frame(frame_izq, bg=C_BORDER2, height=1).pack(fill="x", padx=8)
        (self.f_emp, self.lbl_k_emp, self.lbl_k_emp_e,
                     self.lbl_cf_emp, self.lbl_cf_emp_e) = _hacer_recuadro("──  EMPUJE  ──")

        # Stats mini
        tk.Frame(frame_izq, bg=C_BORDER, height=1).pack(fill="x", pady=(8,0))
        self.lbl_stats = tk.Label(frame_izq, text="", bg="#ffffff", fg=C_TEXT_DIM,
                                   font=(C_MONO,7), justify="left", wraplength=180,
                                   anchor="w", padx=10, pady=6)
        self.lbl_stats.pack(fill="x")

        # ── Panel derecho — Treeview ────────────────────────────────────
        frame_der = tk.Frame(frame_main, bg=C_BG)
        frame_der.pack(side="left", fill="both", expand=True)

        # Sub-header
        frame_tabla_hdr = tk.Frame(frame_der, bg=C_SURFACE, height=36)
        frame_tabla_hdr.pack(fill="x"); frame_tabla_hdr.pack_propagate(False)
        tk.Label(frame_tabla_hdr, text="CICLOS DEL TURNO",
                 bg=C_SURFACE, fg=C_TEXT, font=(C_MONO,9,"bold")).pack(side="left", padx=12)
        self.lbl_ultimo = tk.Label(frame_tabla_hdr, text="",
                                    bg=C_SURFACE, fg=C_TEXT_SUB, font=(C_MONO,8))
        self.lbl_ultimo.pack(side="left", padx=6)
        tk.Frame(frame_der, bg=C_BORDER, height=1).pack(fill="x")

        # Grafico señal en vivo (matplotlib embebido)
        self._buf_vivo = collections.deque(maxlen=int(FS*3))
        self.fig_vivo  = plt.figure(figsize=(9, 1.8), facecolor="#f0f2f5")
        self.fig_vivo.subplots_adjust(left=0.05, right=0.98, top=0.85, bottom=0.22)
        self.ax_vivo   = self.fig_vivo.add_subplot(111)
        self.ax_vivo.set_facecolor("#ffffff")
        self.ax_vivo.tick_params(colors=C_TEXT_SUB, labelsize=6)
        for sp in self.ax_vivo.spines.values(): sp.set_color(C_BORDER2)
        self.ax_vivo.grid(True, alpha=0.08, color=C_BORDER2)
        self.ax_vivo.set_title("Señal en vivo  (últimos 3s)",
                                fontsize=7, color=C_TEXT_SUB, pad=3)
        self.ax_vivo.set_xlabel("Tiempo (s)", fontsize=6, color=C_TEXT_SUB)
        self.canvas_vivo = FigureCanvasTkAgg(self.fig_vivo, master=frame_der)
        self.canvas_vivo.get_tk_widget().pack(fill="x", padx=0)
        tk.Frame(frame_der, bg=C_BORDER, height=1).pack(fill="x")

        # Animación señal en vivo — idéntica a v40
        self._anim = FuncAnimation(self.fig_vivo, self._actualizar_senal,
                                    interval=100, blit=False, cache_frame_data=False)

        # Treeview
        frame_tv = tk.Frame(frame_der, bg=C_BG)
        frame_tv.pack(fill="both", expand=True, padx=6, pady=6)

        style = ttk.Style(); style.theme_use("default")
        style.configure("DT.Treeview",
                         background="#ffffff", foreground="#1a1d27",
                         fieldbackground="#ffffff", rowheight=22,
                         font=(C_MONO,8), relief="flat", bd=0)
        style.configure("DT.Treeview.Heading",
                         background="#e8eaed", foreground="#1a1d27",
                         font=(C_MONO,8,"bold"), relief="flat")
        style.map("DT.Treeview",
                  background=[("selected",C_ACENTO)],
                  foreground=[("selected","white")])

        self.tree = ttk.Treeview(frame_tv, columns=self.COLS,
                                  show="headings", style="DT.Treeview")
        for col, w in zip(self.COLS, self.WIDTHS):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=w, anchor="center", minwidth=36)

        self.tree.tag_configure("MALO",    foreground=C_MALO,    background="#fee2e2")
        self.tree.tag_configure("REVISAR", foreground=C_REVISAR, background="#fef9c3")
        self.tree.tag_configure("BUENO",   foreground=C_BUENO,   background="#dcfce7")
        self.tree.tag_configure("OTRO",    foreground=C_TEXT_SUB)

        scr_y = tk.Scrollbar(frame_tv, orient="vertical", command=self.tree.yview,
                              bg=C_SURFACE2, troughcolor=C_SURFACE, width=5, relief="flat")
        scr_x = tk.Scrollbar(frame_tv, orient="horizontal", command=self.tree.xview,
                              bg=C_SURFACE2, troughcolor=C_SURFACE, width=5, relief="flat")
        self.tree.configure(yscrollcommand=scr_y.set, xscrollcommand=scr_x.set)
        scr_y.pack(side="right", fill="y")
        scr_x.pack(side="bottom", fill="x")
        self.tree.pack(fill="both", expand=True)
        # Doble clic → ventana de segmentos
        self.tree.bind("<Double-1>", self._on_doble_clic)
        # Click derecho → menú contextual
        self.tree.bind("<Button-3>", self._menu_contextual)
        tk.Label(frame_tv,
                 text="Doble clic: segmentos  ·  Click derecho: eliminar fila",
                 bg=C_BG, fg=C_TEXT_DIM,
                 font=("Arial", 7)).pack(anchor="e", pady=(2, 0))

        # ── Barra inferior ────────────────────────────────────────────
        tk.Frame(self, bg=C_BORDER, height=1).pack(fill="x")
        frame_bot = tk.Frame(self, bg=C_SURFACE, height=38)
        frame_bot.pack(fill="x"); frame_bot.pack_propagate(False)
        tk.Label(frame_bot,
                 text="Rojo = DESECHAR  ·  Ámbar = REPARAR  ·  Verde = BUENO",
                 bg=C_SURFACE, fg=C_TEXT_DIM, font=(C_MONO, 7)
                 ).pack(side="left", padx=16)
        tk.Button(frame_bot,
                  text="\u2b07  Exportar Dataset",
                  command=self._exportar_dataset,
                  bg=C_ACENTO, fg="white",
                  activebackground="#3a7fe0", activeforeground="white",
                  relief="flat", bd=0, cursor="hand2",
                  font=("Arial", 8, "bold"), padx=14, pady=6
                  ).pack(side="right", padx=12, pady=4)

    def _guardar_ciclo_json(self, vals, tag):
        import json as _json
        try:
            datos = []
            if os.path.isfile(self._ruta_ciclos):
                with open(self._ruta_ciclos, "r", encoding="utf-8") as f:
                    datos = _json.load(f)
            csv_path = vals[11] if len(vals) > 11 else ""
            csv_nombre = os.path.basename(str(csv_path)) if csv_path else ""
            datos.insert(0, {"n": vals[0], "hora": vals[1],
                              "pinon": vals[2], "resultado": vals[3],
                              "K_ret": vals[4], "CF_ret": vals[5],
                              "K_emp": vals[6], "CF_emp": vals[7],
                              "rms_ret": vals[8], "rms_emp": vals[9],
                              "dur_ret": vals[10], "tag": tag,
                              "csv_nombre": csv_nombre,
                              "operador": self.operador,
                              "turno": self.turno})
            with open(self._ruta_ciclos, "w", encoding="utf-8") as f:
                _json.dump(datos, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _cargar_ciclos_persistidos(self):
        import json as _json
        if not os.path.isfile(self._ruta_ciclos):
            return
        try:
            with open(self._ruta_ciclos, "r", encoding="utf-8") as f:
                datos = _json.load(f)
            if not datos:
                return
            for item in self.tree.get_children():
                self.tree.delete(item)
            self.ciclos.clear(); self._n = 0
            for d in reversed(datos):
                vals = [d["n"], d["hora"], d["pinon"], d["resultado"],
                        d["K_ret"], d["CF_ret"], d["K_emp"], d["CF_emp"],
                        d["rms_ret"], d["rms_emp"], d["dur_ret"]]
                self.tree.insert("", 0, values=vals, tags=(d["tag"],))

                # Reconstruir campos necesarios para doble clic y dashboard
                resultado = d.get("resultado", "BUENO")
                d["golpe"]   = resultado == "REPARAR"
                d["sireneo"] = resultado == "DESECHAR"

                # Reconstruir ruta CSV portátil desde csv_nombre
                # Busca el archivo en las carpetas del programa
                # sin depender del usuario hardcodeado
                if not d.get("csv") or not os.path.isfile(str(d.get("csv",""))):
                    csv_nom = d.get("csv_nombre","")
                    if not csv_nom and d.get("csv"):
                        csv_nom = os.path.basename(str(d["csv"]))
                    if csv_nom:
                        _base = (os.path.dirname(sys.executable)
                                 if getattr(sys, "frozen", False)
                                 else os.path.dirname(os.path.abspath(__file__)))
                        for _sub in ["PIMA_DATOS_ML", "AS14_DATOS_ML",
                                     "AS_DATOS_ML", "."]:
                            _c = os.path.join(_base, _sub, csv_nom)
                            if os.path.isfile(_c):
                                d["csv"] = _c
                                break

                self.ciclos.append(d)
                self._n = max(self._n, int(d["n"]))
            nb   = sum(1 for d in datos if d["resultado"] == "BUENO")
            nrep = sum(1 for d in datos if d["resultado"] == "REPARAR")
            ndes = sum(1 for d in datos if d["resultado"] == "DESECHAR")
            self.lbl_nb.config(text=f"BUENO    {nb}")
            self.lbl_nm.config(text=f"REPARAR  {nrep}")
            self.lbl_nr.config(text=f"DESECHAR {ndes}")
            self.lbl_info_hdr.config(
                text=f"Historial: {len(datos)} ciclos  \u00b7  "
                     f"Op: {self.operador}  \u00b7  Turno: {self.turno}")
        except Exception:
            pass

    def _menu_contextual(self, event):
        iid = self.tree.identify_row(event.y)
        if iid:
            self.tree.selection_set(iid)
        menu = tk.Menu(self, tearoff=0,
                       bg=C_SURFACE, fg=C_TEXT,
                       activebackground=C_MALO, activeforeground="white",
                       relief="flat", bd=0, font=("Arial", 9))
        menu.add_command(label="   Eliminar fila seleccionada",
                         command=self._eliminar_fila)
        menu.add_separator()
        menu.add_command(label="   Eliminar TODO el historial",
                         command=self._eliminar_todo)
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    def _eliminar_fila(self):
        import json as _json
        sel = self.tree.selection()
        if not sel:
            return
        ns = set()
        for iid in sel:
            v = self.tree.item(iid, "values")
            if v:
                try: ns.add(int(v[0]))
                except Exception: pass
            self.tree.delete(iid)
        try:
            if os.path.isfile(self._ruta_ciclos):
                with open(self._ruta_ciclos, "r", encoding="utf-8") as f:
                    datos = _json.load(f)
                datos = [d for d in datos if int(d.get("n", -1)) not in ns]
                with open(self._ruta_ciclos, "w", encoding="utf-8") as f:
                    _json.dump(datos, f, ensure_ascii=False, indent=2)
                nb   = sum(1 for d in datos if d["resultado"] == "BUENO")
                nrep = sum(1 for d in datos if d["resultado"] == "REPARAR")
                ndes = sum(1 for d in datos if d["resultado"] == "DESECHAR")
                self.lbl_nb.config(text=f"BUENO    {nb}")
                self.lbl_nm.config(text=f"REPARAR  {nrep}")
                self.lbl_nr.config(text=f"DESECHAR {ndes}")
                self.lbl_info_hdr.config(text=f"Historial: {len(datos)} ciclos")
        except Exception:
            pass

    def _eliminar_todo(self):
        import json as _json
        if not messagebox.askyesno("Confirmar",
                "Eliminar TODO el historial?" + chr(10) +
                "Esta acción no se puede deshacer.", icon="warning"):
            return
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.ciclos.clear(); self._n = 0
        try:
            with open(self._ruta_ciclos, "w", encoding="utf-8") as f:
                _json.dump([], f)
        except Exception:
            pass
        self.lbl_nb.config(text="BUENO    0")
        self.lbl_nm.config(text="REPARAR  0")
        self.lbl_nr.config(text="DESECHAR 0")
        self.lbl_info_hdr.config(text="Historial vacío")

    def _exportar_dataset(self):
        if not self.ciclos:
            messagebox.showinfo("Sin datos", "No hay ciclos para exportar.")
            return
        fecha_hoy   = datetime.datetime.now().strftime("%d-%m-%Y")
        # Carpeta destino fija: DATOS_MACHINE_LEARNING\DIV-33_YYYY-MM-DD
        _ml_base    = os.path.join(os.path.expanduser("~"),
                                   "OneDrive", "Desktop",
                                   "DATOS_MACHINE_LEARNING")
        carpeta_dia = os.path.join(_ml_base, f"DIV-33_{fecha_hoy}")
        os.makedirs(carpeta_dia, exist_ok=True)
        ts     = datetime.datetime.now().strftime("%H%M%S")
        nombre = f"dataset_{self.operador}_{self.turno}_{ts}.xlsx"
        ruta   = os.path.join(carpeta_dia, nombre)
        try:
            exportar_excel(self.ciclos, self.operador, self.turno, ruta)
            messagebox.showinfo("Exportado",
                "Dataset guardado en:" + chr(10) + chr(10) +
                carpeta_dia + chr(10) + nombre)
            try: os.startfile(carpeta_dia)
            except Exception: pass
        except Exception as e:
            messagebox.showerror("Error al exportar", str(e))

    def _on_doble_clic(self, event):
        """Abre ventana con segmentos de vibración del ciclo seleccionado."""
        item = self.tree.identify_row(event.y)
        if not item:
            return
        # Obtener N° de ciclo desde la primera columna
        vals = self.tree.item(item, "values")
        if not vals:
            return
        try:
            n_ciclo = int(vals[0])
        except (ValueError, IndexError):
            return

        # Buscar el ciclo por campo "n" (robusto independiente del orden)
        ciclo = next((c for c in self.ciclos
                      if str(c.get("n", "")) == str(n_ciclo)), None)
        if ciclo is None:
            ciclo = next((c for i, c in enumerate(self.ciclos, 1)
                          if i == n_ciclo), None)
        if ciclo is None:
            return

        seg_ret = ciclo.get("seg_ret")
        seg_emp = ciclo.get("seg_emp")

        if seg_ret is None and seg_emp is None:
            # Obtener nombre del archivo CSV (sin ruta — portable entre versiones y PCs)
            csv_ruta  = str(ciclo.get("csv") or ciclo.get("ruta_csv", ""))
            csv_nom   = (ciclo.get("csv_nombre") or
                         os.path.basename(csv_ruta) if csv_ruta else "")

            # Construir lista de carpetas donde buscar el CSV
            # Incluye: carpeta del exe/script, carpeta de carpeta_salida del motor,
            # rutas conocidas del proyecto, y la carpeta del usuario actual
            _exe_dir = (os.path.dirname(sys.executable)
                        if getattr(sys, "frozen", False)
                        else os.path.dirname(os.path.abspath(__file__)))
            _horse   = os.path.join(os.path.expanduser("~"),
                                    "OneDrive", "Desktop", "HORSE")
            _desktop = os.path.join(os.path.expanduser("~"), "Desktop", "HORSE")

            _bases = [_exe_dir, _horse, _desktop]
            # Agregar carpeta padre del CSV guardado (aunque sea de otra versión)
            if csv_ruta:
                _bases.append(os.path.dirname(csv_ruta))

            _subcarpetas = ["PIMA_DATOS_ML", "AS14_DATOS_ML",
                            "AS_DATOS_ML", "PIMA_DATOS_ML",
                            "AS15_DATOS_ML", "."]

            csv_encontrado = ""
            # 1. Intentar ruta exacta
            if csv_ruta and os.path.isfile(csv_ruta):
                csv_encontrado = csv_ruta
            # 2. Buscar por nombre en todas las combinaciones base/subcarpeta
            if not csv_encontrado and csv_nom:
                for _base in _bases:
                    for _sub in _subcarpetas:
                        _c = os.path.join(_base, _sub, csv_nom)
                        if os.path.isfile(_c):
                            csv_encontrado = _c
                            ciclo["csv"] = _c   # actualizar para futuras búsquedas
                            break
                    if csv_encontrado:
                        break
                # 3. Búsqueda recursiva en _exe_dir como último recurso
                if not csv_encontrado:
                    for root, dirs, files in os.walk(_exe_dir):
                        if csv_nom in files:
                            csv_encontrado = os.path.join(root, csv_nom)
                            ciclo["csv"] = csv_encontrado
                            break

            if csv_encontrado:
                try:
                    import pandas as _pd
                    df    = _pd.read_csv(csv_encontrado)
                    senal = df["senal"].values.astype(np.float32)
                    feats = analizar_ciclo(senal)
                    if feats:
                        seg_ret = feats.get("seg_ret")
                        seg_emp = feats.get("seg_emp")
                except Exception as _e:
                    messagebox.showerror("Error CSV",
                        "No se pudo re-segmentar: " + str(_e))
                    return
        if seg_ret is None and seg_emp is None:
            csv_nom = ciclo.get("csv_nombre", ciclo.get("csv", "no disponible"))
            messagebox.showinfo("Sin segmentos",
                "Los segmentos no estan disponibles.\n\n"
                "Archivo buscado: " + str(csv_nom))
            return

        self._ventana_segmentos(ciclo, seg_ret, seg_emp, n_ciclo)

    def _ventana_segmentos(self, ciclo, seg_ret, seg_emp, n_ciclo):
        """Ventana con gráficos de los segmentos retroceso y empuje."""
        win = tk.Toplevel(self)
        win.title(f"MAT IA — Ciclo {n_ciclo}  |  {ciclo.get('hora','')}"
                  f"  |  {ciclo.get('pinon','')}")
        win.geometry("900x620")
        win.configure(bg=C_BG)
        _aplicar_icono(win)
        # NO usar transient para no perder foco ni abrir el panel operador

        # Header
        hdr = tk.Frame(win, bg=C_SURFACE, height=44)
        hdr.pack(fill="x"); hdr.pack_propagate(False)
        tk.Frame(hdr, bg=C_ACENTO, width=4).pack(side="left", fill="y")
        fh = tk.Frame(hdr, bg=C_SURFACE); fh.pack(side="left", padx=12)

        golpe   = ciclo.get("golpe",   False)
        sireneo = ciclo.get("sireneo", False)
        if sireneo:   res_txt, res_col = "DESECHAR", C_MALO
        elif golpe:   res_txt, res_col = "REPARAR",  C_REVISAR
        else:         res_txt, res_col = "BUENO",    C_BUENO

        tk.Label(fh,
                 text=f"Ciclo {n_ciclo}  ·  {ciclo.get('pinon','')}  ·  "
                      f"{ciclo.get('hora','')}",
                 bg=C_SURFACE, fg=C_TEXT,
                 font=("Arial", 10, "bold")).pack(anchor="w")
        tk.Label(fh,
                 text=f"Resultado: {res_txt}  ·  "
                      f"K ret={ciclo.get('K_ret','')}  "
                      f"CF ret={ciclo.get('CF_ret','')}  ·  "
                      f"K emp={ciclo.get('K_emp','')}  "
                      f"CF emp={ciclo.get('CF_emp','')}",
                 bg=C_SURFACE, fg=res_col,
                 font=("Arial", 8)).pack(anchor="w")
        tk.Frame(win, bg=C_BORDER, height=1).pack(fill="x")

        # Figura con 2 subplots — usar Figure() directamente para evitar
        # conflictos con el estado global de pyplot
        fig = plt.Figure(figsize=(9, 5.5), facecolor=C_BG)
        fig.subplots_adjust(left=0.07, right=0.97,
                            top=0.92, bottom=0.1, hspace=0.45)
        ax0 = fig.add_subplot(2, 1, 1)
        ax1 = fig.add_subplot(2, 1, 2)

        umb = UMBRALES.get(ciclo.get("pinon", "ARBOL_SEC_14"),
                           UMBRALES["ARBOL_SEC_14"])

        for ax, seg, titulo, k_val, cf_val, k_umb, cf_umb, color in [
            (ax0, seg_ret, "Retroceso",
             ciclo.get("K_ret", ""), ciclo.get("CF_ret", ""),
             umb["K_ret"], umb["CF_ret"], "#1a5fa8"),
            (ax1, seg_emp, "Empuje",
             ciclo.get("K_emp", ""), ciclo.get("CF_emp", ""),
             umb["K_emp"], umb["CF_emp"], "#16a34a"),
        ]:
            ax.set_facecolor("#ffffff")
            ax.tick_params(colors=C_TEXT_DIM, labelsize=6)
            for sp in ax.spines.values(): sp.set_color("#cbd5e1")
            ax.grid(True, alpha=0.15, color="#cbd5e1")
            ax.set_xlabel("Tiempo (s)", fontsize=7, color=C_TEXT_DIM)
            ax.set_ylabel("Amplitud",   fontsize=7, color=C_TEXT_DIM)

            if seg is None or len(seg) == 0:
                ax.text(0.5, 0.5, "Sin datos", ha="center", va="center",
                        transform=ax.transAxes, color=C_TEXT_DIM, fontsize=10)
                ax.set_title(titulo, fontsize=9, color=C_TEXT, pad=4)
                continue

            t = np.arange(len(seg)) / FS
            ax.plot(t, seg, color=color, lw=0.6, alpha=0.85)

            try:
                k_ok  = float(k_val)  <= k_umb
                cf_ok = float(cf_val) <= cf_umb
            except (ValueError, TypeError):
                k_ok = cf_ok = True

            ax.set_title(
                f"{titulo}     K = {k_val}  (umb {k_umb})     "
                f"CF.p99 = {cf_val}  (umb {cf_umb})",
                fontsize=8, color=C_TEXT, pad=4)

            rms_val = float(np.sqrt(np.mean(
                np.array(seg, dtype=np.float64)**2)))
            ax.axhline( rms_val, color="#94a3b8", lw=0.8,
                        ls="--", alpha=0.6, label=f"RMS={rms_val:.4f}")
            ax.axhline(-rms_val, color="#94a3b8", lw=0.8,
                        ls="--", alpha=0.6)
            ax.legend(fontsize=6, loc="upper right",
                      framealpha=0.7, edgecolor="#cbd5e1")

        canvas = FigureCanvasTkAgg(fig, master=win)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)

        # Cerrar la figura al destruir la ventana para liberar memoria
        win.protocol("WM_DELETE_WINDOW",
                     lambda: (plt.close(fig), win.destroy()))

        # Barra inferior
        tk.Frame(win, bg=C_BORDER, height=1).pack(fill="x")
        bf = tk.Frame(win, bg=C_SURFACE, height=32)
        bf.pack(fill="x"); bf.pack_propagate(False)
        tk.Label(bf,
                 text=f"CSV: {ciclo.get('csv', 'No disponible')}",
                 bg=C_SURFACE, fg=C_TEXT_DIM,
                 font=("Arial", 7),
                 anchor="w").pack(side="left", padx=10)
        tk.Button(bf, text="Cerrar",
                  command=lambda: (plt.close(fig), win.destroy()),
                  bg=C_SURFACE2, fg=C_TEXT_DIM,
                  relief="flat", bd=0,
                  font=("Arial", 8),
                  cursor="hand2", padx=12, pady=4
                  ).pack(side="right", padx=8, pady=4)

    def agregar_ciclo(self, ciclo):
        self._n += 1; self.ciclos.append(ciclo)

        # ── Diagnóstico final basado en GOLPE y SIRENEO ──────────────
        # GOLPE   → diente dañado, se puede limar → REPARAR
        # SIRENEO → mala geometría irrecuperable  → DESECHAR
        # Ambos   → SIRENEO es determinante       → DESECHAR
        golpe   = ciclo.get("golpe",   False)
        sireneo = ciclo.get("sireneo", False)
        if sireneo:
            resultado = "DESECHAR"
            tag_color = "MALO"
        elif golpe:
            resultado = "REPARAR"
            tag_color = "REVISAR"
        else:
            resultado = "BUENO"
            tag_color = "BUENO"

        vals = [self._n, ciclo["hora"], ciclo["pinon"], resultado,
                ciclo["K_ret"],  ciclo["CF_ret"],
                ciclo["K_emp"],  ciclo["CF_emp"],
                ciclo["rms_ret"],ciclo["rms_emp"], ciclo["dur_ret"],
                ciclo.get("csv", "")]   # vals[11] = ruta CSV para persistencia
        self.tree.insert("", 0, values=vals[:11], tags=(tag_color,))
        self._guardar_ciclo_json(vals, tag_color)

        # Contadores
        nb = sum(1 for c in self.ciclos
                 if not c.get("golpe") and not c.get("sireneo"))
        nrep = sum(1 for c in self.ciclos
                   if c.get("golpe") and not c.get("sireneo"))
        ndes = sum(1 for c in self.ciclos if c.get("sireneo"))
        self.lbl_nb.config(text=f"BUENO    {nb}")
        self.lbl_nm.config(text=f"REPARAR  {nrep}")
        self.lbl_nr.config(text=f"DESECHAR {ndes}")
        self.lbl_info_hdr.config(
            text=f"Operador: {self.operador}  ·  Turno: {self.turno}  ·  {self._n} ciclos")

        # Actualizar recuadros diagnóstico
        self._actualizar_diag(ciclo)

        # Stats mini
        self.lbl_ultimo.config(
            text=f"{ciclo['hora']}  ·  {resultado}  ·  "
                 f"K={ciclo['K_ret']}  CF={ciclo['CF_ret']}")
        self.lbl_stats.config(
            text=f"Total    : {self._n}\n"
                 f"BUENO    : {nb}\n"
                 f"REPARAR  : {nrep}\n"
                 f"DESECHAR : {ndes}")

    def _actualizar_diag(self, ciclo):
        """
        Colorea los recuadros RETROCESO y EMPUJE.
        Regla: un valor se marca en ROJO solo cuando AMBOS (K y CF)
        superan sus umbrales en esa zona (lógica AND, igual que clasificación).
        Si solo uno supera el umbral → amarillo (REVISAR).
        Si ninguno supera → verde (OK).
        """
        pinon = ciclo.get("pinon", "ARBOL_SEC_14")
        umb   = UMBRALES.get(pinon, UMBRALES["ARBOL_SEC_14"])
        ku_r  = umb["K_ret"];  ku_e  = umb["K_emp"]
        cfu_r = umb["CF_ret"]; cfu_e = umb["CF_emp"]

        kr = ciclo["K_ret"];  cr = ciclo["CF_ret"]
        ke = ciclo["K_emp"];  ce = ciclo["CF_emp"]

        malo_ret = (kr > ku_r) and (cr > cfu_r)
        malo_emp = (ke > ku_e) and (ce > cfu_e)

        sosp_ret = (kr > ku_r) != (cr > cfu_r)
        sosp_emp = (ke > ku_e) != (ce > cfu_e)

        def color_zona(malo, sosp):
            if malo: return C_MALO
            if sosp: return C_REVISAR
            return C_BUENO

        def color_val(v, umbral, zona_malo):
            if zona_malo and v > umbral: return C_MALO
            if v > umbral:               return C_REVISAR
            return C_TEXT

        def etiq_k_ret(v):
            if v > ku_r: return "[SOBRE UMB]"
            return "[OK]"

        def etiq_k_emp(v):
            if v > ku_e: return "[SOBRE UMB]"
            return "[OK]"

        def etiq_cf_ret(v):
            if v > cfu_r: return "[SOBRE UMB]"
            return "[OK]"

        def etiq_cf_emp(v):
            if v > cfu_e: return "[SOBRE UMB]"
            return "[OK]"

        def borde(malo, sosp):
            if malo: return C_MALO
            if sosp: return C_REVISAR
            return C_BUENO

        # ── Retroceso ──
        col_kr = color_val(kr, ku_r,  malo_ret)
        col_cr = color_val(cr, cfu_r, malo_ret)
        self.lbl_k_ret.config(  text=f"{kr:.2f}", fg=col_kr)
        self.lbl_k_ret_e.config(text=etiq_k_ret(kr),  fg=col_kr)
        self.lbl_cf_ret.config( text=f"{cr:.2f}", fg=col_cr)
        self.lbl_cf_ret_e.config(text=etiq_cf_ret(cr), fg=col_cr)
        self.f_ret.config(highlightbackground=borde(malo_ret, sosp_ret),
                          highlightthickness=2)

        # ── Empuje ──
        col_ke = color_val(ke, ku_e,  malo_emp)
        col_ce = color_val(ce, cfu_e, malo_emp)
        self.lbl_k_emp.config(  text=f"{ke:.2f}", fg=col_ke)
        self.lbl_k_emp_e.config(text=etiq_k_emp(ke),  fg=col_ke)
        self.lbl_cf_emp.config( text=f"{ce:.2f}", fg=col_ce)
        self.lbl_cf_emp_e.config(text=etiq_cf_emp(ce), fg=col_ce)
        self.f_emp.config(highlightbackground=borde(malo_emp, sosp_emp),
                          highlightthickness=2)

    def push_vivo(self, datos):
        """Recibe muestras del motor para el grafico en vivo."""
        self._buf_vivo.extend(datos.tolist())

    def _actualizar_senal(self, frame):
        ax = self.ax_vivo; ax.clear()
        ax.set_facecolor("#ffffff")
        ax.tick_params(colors=C_TEXT_SUB, labelsize=6)
        for sp in ax.spines.values(): sp.set_color(C_BORDER2)
        ax.grid(True, alpha=0.08, color=C_BORDER2)
        buf = list(self._buf_vivo)
        if buf:
            arr = np.array(buf, dtype=np.float32)
            t   = np.arange(len(arr)) / FS
            ax.plot(t, arr, color=C_ACENTO, lw=0.5, alpha=0.9)
            ax.set_xlim(0, max(t[-1], 3))
        else:
            ax.set_xlim(0, 3)
            ax.text(1.5, 0, "Sin señal — esperando dispositivo",
                    ha="center", color=C_TEXT_DIM, fontsize=7,
                    transform=ax.transData)
        ax.set_title("Señal en vivo  (últimos 3s)", fontsize=7,
                     color=C_TEXT_SUB, pad=3)
        ax.set_xlabel("Tiempo (s)", fontsize=6, color=C_TEXT_SUB)

    def _limpiar(self):
        if messagebox.askyesno("Limpiar", "¿Borrar todos los registros de la tabla?"):
            self.tree.delete(*self.tree.get_children())
            self.ciclos.clear(); self._n = 0
            self.lbl_nb.config(text="BUENO   0")
            self.lbl_nm.config(text="MALO    0")
            self.lbl_nr.config(text="REVISAR 0")
            self.lbl_stats.config(text="")
            self.lbl_ultimo.config(text="")

    def _exportar(self):
        if not self.ciclos:
            messagebox.showinfo("Sin datos","No hay ciclos registrados."); return
        fecha  = datetime.datetime.now().strftime("%Y%m%d_%H%M")
        nombre = f"turno_{self.operador.replace(' ','_')}_{fecha}.xlsx"
        ruta = filedialog.asksaveasfilename(title="Guardar historial del turno",
                                             defaultextension=".xlsx",
                                             initialfile=nombre,
                                             filetypes=[("Excel","*.xlsx")])
        if not ruta: return
        try:
            exportar_excel(self.ciclos, self.operador, self.turno, ruta)
            self.lbl_export.config(text=f"✓ Guardado: {os.path.basename(ruta)}")
            if sys.platform=="win32": os.startfile(os.path.dirname(ruta))
        except Exception as e:
            messagebox.showerror("Error al exportar", str(e))

    def exportar_automatico(self, carpeta_salida):
        """Guarda la tabla al cerrar. Retorna ruta guardada o lanza excepción."""
        if not self.ciclos:
            return None
        fecha_hoy  = datetime.datetime.now().strftime("%d-%m-%Y")
        fecha_hora = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        carpeta_dia = os.path.join(carpeta_salida, f"DIV-33_{fecha_hoy}")
        os.makedirs(carpeta_dia, exist_ok=True)
        nombre = (f"turno_{self.operador.replace(' ','_')}_"
                  f"{self.turno}_{fecha_hora}.xlsx")
        ruta = os.path.join(carpeta_dia, nombre)
        exportar_excel(self.ciclos, self.operador, self.turno, ruta)
        return ruta

# ═══════════════════════════════════════════════════════════════════════════
# VENTANA 1 — PANEL OPERADOR
# ═══════════════════════════════════════════════════════════════════════════


class VentanaDashboard(tk.Toplevel):
    """
    Cuarta pestaña — Dashboard en tiempo real.
    Lee ciclos_turno.json al abrir (misma fuente que la tabla).
    Se actualiza con cada pieza nueva durante la sesión.
    """

    def __init__(self, master):
        super().__init__(master)
        self.title("MAT IA — Dashboard")
        self.geometry("1100x700"); self.minsize(900, 580)
        self.configure(bg=C_BG)
        _aplicar_icono(self)
        self.ciclos = []
        # Misma ruta JSON que VentanaTurno
        _base = (os.path.dirname(sys.executable)
                 if getattr(sys, "frozen", False)
                 else os.path.dirname(os.path.abspath(__file__)))
        self._ruta_ciclos = os.path.join(_base, "ciclos_turno.json")
        self._build()
        # Cargar historial persistido al abrir
        self._cargar_desde_json()

    def _build(self):
        # ── Header ───────────────────────────────────────────────────────
        hdr = tk.Frame(self, bg=C_SURFACE, height=56)
        hdr.pack(fill="x"); hdr.pack_propagate(False)
        tk.Frame(hdr, bg=C_ACENTO, width=4).pack(side="left", fill="y")
        _logo = _cargar_logo()
        if _logo:
            tk.Label(hdr, image=_logo, bg=C_SURFACE
                     ).pack(side="left", padx=14, pady=8)
        else:
            tk.Label(hdr, text="HORSE", bg=C_SURFACE, fg=C_ACENTO,
                     font=(C_MONO, 13, "bold")).pack(side="left", padx=14)
        tk.Frame(hdr, bg=C_BORDER, width=1).pack(side="left", fill="y", pady=8)
        fh = tk.Frame(hdr, bg=C_SURFACE); fh.pack(side="left", padx=12)
        tk.Label(fh, text="DASHBOARD DE CALIDAD",
                 bg=C_SURFACE, fg=C_TEXT,
                 font=("Arial", 12, "bold")).pack(anchor="w")
        self.lbl_sub_hdr = tk.Label(fh, text="Sin datos aún",
                                     bg=C_SURFACE, fg=C_TEXT_SUB,
                                     font=("Arial", 9))
        self.lbl_sub_hdr.pack(anchor="w")
        _lm = _hacer_logo_matia(hdr, bg=C_SURFACE)
        _lm.pack(side="right", padx=14, pady=6)
        tk.Frame(self, bg=C_BORDER, height=1).pack(fill="x")

        # ── Tarjetas métricas superiores ─────────────────────────────────
        f_cards = tk.Frame(self, bg=C_BG)
        f_cards.pack(fill="x", padx=16, pady=(10, 4))
        for i in range(5):
            f_cards.columnconfigure(i, weight=1)

        def card(parent, col, titulo, var_texto, color):
            f = tk.Frame(parent, bg=C_SURFACE2,
                         highlightbackground=C_BORDER,
                         highlightthickness=1)
            f.grid(row=0, column=col, sticky="ew",
                   padx=(0 if col == 0 else 6, 0), ipady=6)
            tk.Label(f, text=titulo, bg=C_SURFACE2, fg=C_TEXT_SUB,
                     font=("Arial", 8)).pack(anchor="w", padx=10, pady=(6,0))
            lbl = tk.Label(f, textvariable=var_texto,
                           bg=C_SURFACE2, fg=color,
                           font=("Arial", 18, "bold"))
            lbl.pack(anchor="w", padx=10, pady=(0, 6))
            return lbl

        self._v_total   = tk.StringVar(value="0")
        self._v_bueno   = tk.StringVar(value="0")
        self._v_golpe   = tk.StringVar(value="0")
        self._v_sireneo = tk.StringVar(value="0")
        self._v_tasa    = tk.StringVar(value="0%")

        card(f_cards, 0, "TOTAL PIEZAS",    self._v_total,   C_TEXT)
        card(f_cards, 1, "BUENAS",          self._v_bueno,   C_BUENO)
        card(f_cards, 2, "CON GOLPE",       self._v_golpe,   C_REVISAR)
        card(f_cards, 3, "SIRENEO",         self._v_sireneo, C_MALO)
        card(f_cards, 4, "TASA DEFECTOS",   self._v_tasa,    C_MALO)

        # ── Figura matplotlib con 6 subplots ─────────────────────────────
        self.fig = plt.Figure(figsize=(13, 7), facecolor=C_BG)
        self.fig.subplots_adjust(
            left=0.06, right=0.97,
            top=0.93, bottom=0.10,
            hspace=0.55, wspace=0.38)

        # Disposición: 2 filas × 3 columnas
        self.ax_donut  = self.fig.add_subplot(2, 3, 1)
        self.ax_giro   = self.fig.add_subplot(2, 3, 2)
        self.ax_pinon  = self.fig.add_subplot(2, 3, 3)
        self.ax_kvscf  = self.fig.add_subplot(2, 3, 4)
        self.ax_evol   = self.fig.add_subplot(2, 3, 5)
        self.ax_tasa   = self.fig.add_subplot(2, 3, 6)

        for ax in [self.ax_donut, self.ax_giro, self.ax_pinon,
                   self.ax_kvscf, self.ax_evol, self.ax_tasa]:
            ax.set_facecolor(C_SURFACE)
            ax.tick_params(colors=C_TEXT_SUB, labelsize=7)
            for sp in ax.spines.values():
                sp.set_color(C_BORDER)

        # ── Barra inferior — exportar PDF ─────────────────────────────
        tk.Frame(self, bg=C_BORDER, height=1).pack(fill="x", side="bottom")
        f_bot = tk.Frame(self, bg=C_SURFACE, height=42)
        f_bot.pack(fill="x", side="bottom"); f_bot.pack_propagate(False)
        tk.Label(f_bot,
                 text="Exportar dashboard como PDF — elige el período:",
                 bg=C_SURFACE, fg=C_TEXT_DIM, font=("Arial", 8)
                 ).pack(side="left", padx=14)
        for txt, periodo in [
            ("📄  Diario",  "dia"),
            ("📄  Semanal", "semana"),
            ("📄  Mensual", "mes"),
        ]:
            tk.Button(f_bot, text=txt,
                      command=lambda p=periodo: self._exportar_pdf(p),
                      bg=C_ACENTO, fg="white",
                      activebackground="#3a7fe0", activeforeground="white",
                      relief="flat", bd=0, cursor="hand2",
                      font=("Arial", 8, "bold"), padx=12, pady=6
                      ).pack(side="right", padx=(0, 8), pady=4)

        self.canvas_fig = FigureCanvasTkAgg(self.fig, master=self)
        self.canvas_fig.get_tk_widget().pack(
            fill="both", expand=True, padx=8, pady=(0, 4))

        self._dibujar_vacios()

    def _exportar_pdf(self, periodo):
        """
        Exporta el dashboard como PDF.
        periodo: "dia" | "semana" | "mes"
        Destino: C:\\Users\\usuario\\OneDrive\\Desktop\\DATOS_MACHINE_LEARNING\\DASHBOARD\\
        Nombres:
          Diario  : DASHBOARD_DIARIO_YYYY-MM-DD.pdf
          Semanal : DASHBOARD_SEMANAL_SEMANA{N}_{YYYY}.pdf
          Mensual : DASHBOARD_MENSUAL_{MES}-{YYYY}.pdf
        """
        if not self.ciclos:
            messagebox.showinfo("Sin datos",
                "No hay datos para exportar el dashboard.")
            return

        import calendar
        ahora   = datetime.datetime.now()
        año     = ahora.strftime("%Y")
        mes_num = ahora.strftime("%m")
        dia_str = ahora.strftime("%Y-%m-%d")
        semana  = ahora.isocalendar()[1]   # semana ISO del año
        mes_nom = ahora.strftime("%B").upper()   # ABRIL, MAYO...

        if periodo == "dia":
            nombre_pdf = f"DASHBOARD_DIARIO_{dia_str}.pdf"
            titulo_rep = f"Dashboard Diario  —  {dia_str}"
        elif periodo == "semana":
            nombre_pdf = f"DASHBOARD_SEMANAL_SEMANA{semana}_{año}.pdf"
            titulo_rep = f"Dashboard Semanal  —  Semana {semana} de {año}"
        else:
            nombre_pdf = f"DASHBOARD_MENSUAL_{mes_nom}-{año}.pdf"
            titulo_rep = f"Dashboard Mensual  —  {mes_nom} {año}"

        # Carpeta destino
        _base_ml  = os.path.join(os.path.expanduser("~"),
                                  "OneDrive", "Desktop",
                                  "DATOS_MACHINE_LEARNING")
        carpeta   = os.path.join(_base_ml, "DASHBOARD")
        os.makedirs(carpeta, exist_ok=True)
        ruta_pdf  = os.path.join(carpeta, nombre_pdf)

        try:
            from matplotlib.backends.backend_pdf import PdfPages
            import matplotlib.pyplot as _plt

            n    = len(self.ciclos)
            cs   = self.ciclos
            buenos  = sum(1 for c in cs if not c.get("golpe") and not c.get("sireneo"))
            golpes  = sum(1 for c in cs if c.get("golpe") and not c.get("sireneo"))
            sirenos = sum(1 for c in cs if c.get("sireneo"))
            tasa    = (golpes + sirenos) / n * 100 if n else 0

            with PdfPages(ruta_pdf) as pdf:
                # ── Página 1: Portada + métricas + 6 gráficos ─────────────
                fig = _plt.figure(figsize=(11.69, 8.27))  # A4 landscape
                fig.patch.set_facecolor("#f0f2f5")

                # Título portada
                fig.text(0.5, 0.97, "MAT IA  —  Sistema NVH DEMM",
                         ha="center", va="top",
                         fontsize=15, fontweight="bold", color="#1a1d27")
                fig.text(0.5, 0.93, titulo_rep,
                         ha="center", va="top",
                         fontsize=11, color="#1a5fa8")
                fig.text(0.5, 0.905,
                         f"Generado: {ahora.strftime("%d/%m/%Y %H:%M")}  ·  "
                         f"Total piezas: {n}  ·  "
                         f"Buenas: {buenos}  ·  "
                         f"Golpe: {golpes}  ·  "
                         f"Sireneo: {sirenos}  ·  "
                         f"Tasa defectos: {tasa:.1f}%",
                         ha="center", va="top",
                         fontsize=8, color="#4a5068")
                fig.text(0.98, 0.01,
                         f"DATOS_MACHINE_LEARNING/DASHBOARD/{nombre_pdf}",
                         ha="right", va="bottom",
                         fontsize=6, color="#8a90a8")

                # 6 subplots copiados de la figura actual
                axes_src = [self.ax_donut, self.ax_giro, self.ax_pinon,
                            self.ax_kvscf, self.ax_evol,  self.ax_tasa]
                positions = [
                    [0.04, 0.08, 0.27, 0.76],
                    [0.36, 0.08, 0.27, 0.76],
                    [0.68, 0.08, 0.27, 0.76],
                ]
                # Fila 1: donut, giro, piñón
                for i, ax_src in enumerate(axes_src[:3]):
                    ax_new = fig.add_axes([
                        positions[i][0],
                        0.50,
                        positions[i][2],
                        0.38,
                    ])
                    self._copiar_ax(ax_src, ax_new)
                # Fila 2: kvscf, evol, tasa
                for i, ax_src in enumerate(axes_src[3:]):
                    ax_new = fig.add_axes([
                        positions[i][0],
                        0.06,
                        positions[i][2],
                        0.38,
                    ])
                    self._copiar_ax(ax_src, ax_new)

                pdf.savefig(fig, bbox_inches="tight")
                _plt.close(fig)

            messagebox.showinfo("PDF exportado",
                "Dashboard guardado en:" + chr(10) + chr(10) +
                carpeta + chr(10) + nombre_pdf)
            try: os.startfile(ruta_pdf)
            except Exception: pass

        except Exception as e:
            messagebox.showerror("Error al exportar PDF", str(e))

    def _copiar_ax(self, ax_src, ax_dst):
        """Redibuja ax_src en ax_dst para el PDF. Maneja Wedge, Rectangle, scatter y líneas."""
        import matplotlib.patches as mpatches
        ax_dst.set_facecolor(ax_src.get_facecolor())
        ax_dst.set_title(ax_src.get_title(), fontsize=7, color="#1a1d27", pad=4)
        # ── Líneas
        for line in ax_src.get_lines():
            ax_dst.plot(line.get_xdata(), line.get_ydata(),
                        color=line.get_color(), linewidth=line.get_linewidth(),
                        linestyle=line.get_linestyle(), marker=line.get_marker(),
                        markersize=line.get_markersize(), alpha=line.get_alpha() or 1.0)
        # ── Patches: Wedge (donut) o Rectangle (barras)
        for patch in ax_src.patches:
            try:
                if isinstance(patch, mpatches.Wedge):
                    w = mpatches.Wedge(
                        patch.center, patch.r, patch.theta1, patch.theta2,
                        width=patch.width,
                        facecolor=patch.get_facecolor(),
                        edgecolor=patch.get_edgecolor(),
                        linewidth=patch.get_linewidth())
                    ax_dst.add_patch(w)
                elif isinstance(patch, mpatches.Rectangle):
                    r = mpatches.Rectangle(
                        patch.get_xy(), patch.get_width(), patch.get_height(),
                        facecolor=patch.get_facecolor(),
                        edgecolor=patch.get_edgecolor(),
                        linewidth=patch.get_linewidth(),
                        alpha=patch.get_alpha() or 1.0)
                    ax_dst.add_patch(r)
            except Exception:
                pass
        # ── Scatter
        for col in ax_src.collections:
            try:
                offsets = col.get_offsets()
                if len(offsets):
                    fc = col.get_facecolor()
                    sz = col.get_sizes()
                    ax_dst.scatter(
                        offsets[:, 0], offsets[:, 1],
                        c=fc if len(fc) == len(offsets) else [fc[0]] * len(offsets),
                        s=sz[0] if len(sz) else 20,
                        alpha=col.get_alpha() or 0.75, edgecolors="none")
            except Exception:
                pass
        # ── Textos
        for txt in ax_src.texts:
            try:
                ax_dst.text(
                    txt.get_position()[0], txt.get_position()[1], txt.get_text(),
                    transform=ax_dst.transAxes
                        if txt.get_transform() == ax_src.transAxes else ax_dst.transData,
                    ha=txt.get_ha(), va=txt.get_va(),
                    color=txt.get_color(), fontsize=7, alpha=txt.get_alpha() or 1.0)
            except Exception:
                pass
        # ── Leyenda
        try:
            handles, labels = ax_src.get_legend_handles_labels()
            if handles:
                ax_dst.legend(handles, labels, fontsize=6, frameon=False,
                              loc="best", labelcolor="#4a5068")
        except Exception:
            pass
        # ── Ejes
        ax_dst.set_xlabel(ax_src.get_xlabel(), fontsize=6, color="#4a5068")
        ax_dst.set_ylabel(ax_src.get_ylabel(), fontsize=6, color="#4a5068")
        ax_dst.tick_params(labelsize=6, colors="#4a5068")
        for sp in ax_dst.spines.values():
            sp.set_color("#c8cdd8")
        try:
            if ax_src.get_xlim() != (0.0, 1.0):
                ax_dst.set_xlim(ax_src.get_xlim())
            if ax_src.get_ylim() != (0.0, 1.0):
                ax_dst.set_ylim(ax_src.get_ylim())
        except Exception:
            pass
        try:
            ticks = ax_src.get_xticks()
            tlbls = [t.get_text() for t in ax_src.get_xticklabels()]
            if any(tlbls):
                ax_dst.set_xticks(ticks)
                ax_dst.set_xticklabels(tlbls, fontsize=6)
        except Exception:
            pass


    def _cargar_desde_json(self):

        """Carga ciclos_turno.json al abrir — sincroniza con la tabla."""
        import json as _json
        if not os.path.isfile(self._ruta_ciclos):
            return
        try:
            with open(self._ruta_ciclos, "r", encoding="utf-8") as f:
                datos = _json.load(f)
            if not datos:
                return
            # Convertir al formato que usa actualizar()
            # El JSON guarda: n, hora, pinon, resultado, K_ret, CF_ret,
            #                 K_emp, CF_emp, rms_ret, rms_emp, dur_ret,
            #                 tag, operador, turno
            # actualizar() usa: golpe, sireneo, zona_golpe, K_ret, CF_ret,
            #                   K_emp, CF_emp, pinon
            self.ciclos = []
            for d in reversed(datos):   # reversed: JSON tiene más reciente primero
                resultado = d.get("resultado", "BUENO")
                self.ciclos.append({
                    "golpe":      resultado == "REPARAR",
                    "sireneo":    resultado == "DESECHAR",
                    "zona_golpe": d.get("zona_golpe", ""),
                    "K_ret":      float(d.get("K_ret",  0)),
                    "CF_ret":     float(d.get("CF_ret", 0)),
                    "K_emp":      float(d.get("K_emp",  0)),
                    "CF_emp":     float(d.get("CF_emp", 0)),
                    "rms_ret":    float(d.get("rms_ret", 0)),
                    "rms_emp":    float(d.get("rms_emp", 0)),
                    "pinon":      d.get("pinon", "ARBOL_SEC_14"),
                    "hora":       d.get("hora",  ""),
                    "operador":   d.get("operador", ""),
                    "turno":      d.get("turno",  ""),
                    "resultado":  resultado,
                })
            self.actualizar()
        except Exception:
            pass

    def agregar_ciclo_dashboard(self, ciclo):
        """
        Llamado desde _on_ciclo en tiempo real.
        Agrega el ciclo a self.ciclos y redibuja.
        El ciclo ya viene en el formato del motor (golpe, sireneo, etc.)
        """
        self.ciclos.append(ciclo)
        self.actualizar()

    def _dibujar_vacios(self):
        """Dibuja gráficos vacíos con títulos al abrir."""
        for ax, titulo in [
            (self.ax_donut, "Distribución BUENO / GOLPE / SIRENEO"),
            (self.ax_giro,  "Golpes por giro"),
            (self.ax_pinon, "Resultados por piñón"),
            (self.ax_kvscf, "Dispersión K vs CF"),
            (self.ax_evol,  "Evolución K y CF en el turno"),
            (self.ax_tasa,  "Tasa de defectos acumulada"),
        ]:
            ax.set_title(titulo, fontsize=8, color=C_TEXT, pad=6)
            ax.text(0.5, 0.5, "Sin datos", transform=ax.transAxes,
                    ha="center", va="center",
                    color=C_TEXT_DIM, fontsize=9)
        self.canvas_fig.draw()

    def actualizar(self):
        """Recalcula y redibuja todos los gráficos con self.ciclos."""
        if not self.winfo_exists():
            return
        cs = self.ciclos
        if not cs:
            return

        n      = len(cs)
        buenos  = sum(1 for c in cs if not c.get("golpe") and not c.get("sireneo"))
        golpes  = sum(1 for c in cs if c.get("golpe") and not c.get("sireneo"))
        sirenos = sum(1 for c in cs if c.get("sireneo"))
        defect  = golpes + sirenos
        tasa    = defect / n * 100 if n else 0

        # Tarjetas
        self._v_total.set(str(n))
        self._v_bueno.set(str(buenos))
        self._v_golpe.set(str(golpes))
        self._v_sireneo.set(str(sirenos))
        self._v_tasa.set(f"{tasa:.1f}%")
        self.lbl_sub_hdr.config(
            text=f"{n} piezas analizadas  ·  "
                 f"{buenos} buenas  ·  "
                 f"{golpes} golpe  ·  "
                 f"{sirenos} sireneo")

        # Limpiar
        for ax in [self.ax_donut, self.ax_giro, self.ax_pinon,
                   self.ax_kvscf, self.ax_evol, self.ax_tasa]:
            ax.clear()
            ax.set_facecolor(C_SURFACE)
            ax.tick_params(colors=C_TEXT_SUB, labelsize=7)
            for sp in ax.spines.values():
                sp.set_color(C_BORDER)

        # ── 1. Donut BUENO / GOLPE / SIRENEO ─────────────────────────────
        ax = self.ax_donut
        vals_d  = [buenos, golpes, sirenos]
        labs_d  = ["BUENO", "GOLPE", "SIRENEO"]
        cols_d  = [C_BUENO, C_REVISAR, C_MALO]
        # Filtrar ceros
        data_d = [(v, l, c) for v, l, c in zip(vals_d, labs_d, cols_d) if v > 0]
        if data_d:
            vd, ld, cd = zip(*data_d)
            wedges, _ = ax.pie(
                vd, colors=cd,
                startangle=90,
                wedgeprops=dict(width=0.55, edgecolor=C_BG, linewidth=2))
            ax.legend(wedges,
                      [f"{l}  {v}" for v, l in zip(vd, ld)],
                      loc="lower center", bbox_to_anchor=(0.5, -0.18),
                      ncol=3, fontsize=7,
                      frameon=False,
                      labelcolor=C_TEXT_SUB)
        ax.set_title("Distribución BUENO / GOLPE / SIRENEO",
                     fontsize=8, color=C_TEXT, pad=6)

        # ── 2. Golpes por giro ────────────────────────────────────────────
        ax = self.ax_giro
        con_golpe = [c for c in cs if c.get("golpe")]
        g_pri = sum(1 for c in con_golpe if c.get("zona_golpe","") == "EMP")
        g_seg = sum(1 for c in con_golpe if c.get("zona_golpe","") == "RET")
        g_amb = sum(1 for c in con_golpe
                    if c.get("zona_golpe","") in ("RET+EMP","EMP+RET"))
        g_unk = len(con_golpe) - g_pri - g_seg - g_amb
        zonas = ["Primer giro", "Segundo giro", "Ambos", "Otro"]
        vals_g = [g_pri, g_seg, g_amb, g_unk]
        cols_g = [C_MALO, C_REVISAR, "#f97316", C_TEXT_DIM]
        idx = [i for i, v in enumerate(vals_g) if v > 0]
        if idx:
            bars = ax.bar([zonas[i] for i in idx],
                          [vals_g[i] for i in idx],
                          color=[cols_g[i] for i in idx],
                          width=0.55, edgecolor=C_BG, linewidth=1.5)
            ax.bar_label(bars, fmt="%d", color=C_TEXT_SUB,
                         fontsize=8, padding=2)
        ax.set_title("Golpes por giro", fontsize=8, color=C_TEXT, pad=6)
        ax.set_ylabel("Piezas", fontsize=7, color=C_TEXT_SUB)
        ax.tick_params(axis="x", labelsize=7)
        if not idx:
            ax.text(0.5, 0.5, "Sin golpes", transform=ax.transAxes,
                    ha="center", va="center",
                    color=C_TEXT_DIM, fontsize=9)

        # ── 3. Resultados por piñón ───────────────────────────────────────
        ax = self.ax_pinon
        pinones = ["PIMA", "ARBOL_SEC_14"]
        nombres = ["PIMA", "Árbol Sec."]
        x = range(len(pinones))
        w = 0.25
        bue_p = [sum(1 for c in cs
                     if c.get("pinon") == p
                     and not c.get("golpe") and not c.get("sireneo"))
                 for p in pinones]
        gol_p = [sum(1 for c in cs
                     if c.get("pinon") == p and c.get("golpe")
                     and not c.get("sireneo"))
                 for p in pinones]
        sir_p = [sum(1 for c in cs
                     if c.get("pinon") == p and c.get("sireneo"))
                 for p in pinones]
        b1 = ax.bar([i - w for i in x], bue_p, w,
                    color=C_BUENO,   label="Bueno",   edgecolor=C_BG)
        b2 = ax.bar([i     for i in x], gol_p, w,
                    color=C_REVISAR, label="Golpe",   edgecolor=C_BG)
        b3 = ax.bar([i + w for i in x], sir_p, w,
                    color=C_MALO,    label="Sireneo", edgecolor=C_BG)
        ax.set_xticks(list(x)); ax.set_xticklabels(nombres, fontsize=7)
        ax.legend(fontsize=7, frameon=False, labelcolor=C_TEXT_SUB,
                  loc="upper right")
        ax.set_title("Resultados por piñón", fontsize=8, color=C_TEXT, pad=6)
        ax.set_ylabel("Piezas", fontsize=7, color=C_TEXT_SUB)

        # ── 4. Dispersión K vs CF ─────────────────────────────────────────
        ax = self.ax_kvscf
        pinon_actual = cs[-1].get("pinon", "ARBOL_SEC_14") if cs else "ARBOL_SEC_14"
        umb = UMBRALES.get(pinon_actual, UMBRALES["ARBOL_SEC_14"])
        k_umb  = umb.get("K_ret",  umb.get("K",  5.0))
        cf_umb = umb.get("CF_ret", umb.get("CF", 3.5))

        def color_punto(c):
            if c.get("sireneo"): return C_MALO
            if c.get("golpe"):   return C_REVISAR
            return C_BUENO

        ks  = [c.get("K_ret",  0) for c in cs]
        cfs = [c.get("CF_ret", 0) for c in cs]
        cols_sc = [color_punto(c) for c in cs]
        ax.scatter(ks, cfs, c=cols_sc, s=28, alpha=0.75, edgecolors="none")
        ax.axvline(k_umb,  color=C_MALO, linewidth=0.8,
                   linestyle="--", alpha=0.7)
        ax.axhline(cf_umb, color=C_MALO, linewidth=0.8,
                   linestyle="--", alpha=0.7)
        ax.set_xlabel("K ret", fontsize=7, color=C_TEXT_SUB)
        ax.set_ylabel("CF ret", fontsize=7, color=C_TEXT_SUB)
        ax.set_title("Dispersión K vs CF (retroceso)",
                     fontsize=8, color=C_TEXT, pad=6)

        # ── 5. Evolución K y CF en el turno ──────────────────────────────
        ax = self.ax_evol
        idx_t = list(range(1, n + 1))
        k_serie  = [c.get("K_ret",  0) for c in cs]
        cf_serie = [c.get("CF_ret", 0) for c in cs]
        ax.plot(idx_t, k_serie,  color=C_ACENTO, linewidth=1.2,
                label="K ret",  marker="o", markersize=3)
        ax.plot(idx_t, cf_serie, color=C_REVISAR, linewidth=1.2,
                label="CF ret", marker="s", markersize=3,
                linestyle="--")
        ax.axhline(k_umb,  color=C_ACENTO, linewidth=0.7,
                   linestyle=":", alpha=0.6)
        ax.axhline(cf_umb, color=C_REVISAR, linewidth=0.7,
                   linestyle=":", alpha=0.6)
        ax.set_xlabel("Pieza N°", fontsize=7, color=C_TEXT_SUB)
        ax.legend(fontsize=7, frameon=False, labelcolor=C_TEXT_SUB)
        ax.set_title("Evolución K y CF en el turno",
                     fontsize=8, color=C_TEXT, pad=6)

        # ── 6. Tasa de defectos acumulada ─────────────────────────────────
        ax = self.ax_tasa
        tasa_acum = []
        n_def = 0
        for i, c in enumerate(cs, 1):
            if c.get("golpe") or c.get("sireneo"):
                n_def += 1
            tasa_acum.append(n_def / i * 100)
        ax.plot(range(1, n + 1), tasa_acum,
                color=C_MALO, linewidth=1.4)
        ax.fill_between(range(1, n + 1), tasa_acum,
                        color=C_MALO, alpha=0.12)
        ax.set_xlabel("Pieza N°", fontsize=7, color=C_TEXT_SUB)
        ax.set_ylabel("%", fontsize=7, color=C_TEXT_SUB)
        ax.set_ylim(0, max(max(tasa_acum) * 1.2, 10))
        ax.set_title("Tasa de defectos acumulada",
                     fontsize=8, color=C_TEXT, pad=6)

        self.canvas_fig.draw()

class VentanaOperador(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("MAT IA — Panel Operador")
        self.geometry("520x560")
        self.minsize(420, 480)
        self.resizable(True, True)
        self.configure(bg=C_BG)
        _aplicar_icono(self)
        self.motor       = Motor()
        self.var_op      = tk.StringVar()
        self.var_turno   = tk.StringVar(value="Mañana")
        self.var_dev     = tk.StringVar()
        self.lbl_id      = tk.StringVar(value="Sin cargar")
        self.lbl_pima    = tk.StringVar(value="Sin cargar")
        self.lbl_as      = tk.StringVar(value="Sin cargar")
        self._activo     = False
        self._vs         = None
        self._vt         = None
        self._vd         = None   # Dashboard
        self._ciclos_turno = []   # copia local de todos los ciclos del turno
        self._devs       = []
        self._dev_id_auto = None
        self.var_carpeta  = None
        # ── Historial de operadores ──────────────────────────────────
        self._ruta_historial = os.path.join(
            os.path.expanduser("~"), "OneDrive", "Desktop", "HORSE",
            "operadores_historial.json")
        self._historial_ops  = self._cargar_historial()
        self._build()
        self._listar_devs()
        # pinon_manual inicia en None — el operador selecciona en el semáforo
        self.motor.pinon_manual = None
        self.after(200, self._cargar_modelos_default)

    def _build(self):
        # ── Inicializar variables internas (no visibles) ──────────────
        _carpeta_default = r"C:\Users\usuario\Desktop\DATOS_MACHINE_LEARNING"
        self.var_carpeta = tk.StringVar(value=_carpeta_default)
        self.motor.carpeta_salida = _carpeta_default
        # Widgets ocultos necesarios para compatibilidad con métodos existentes
        self.lbl_dev_nombre = tk.Label(self, text="")   # actualizado por _listar_devs
        self.f_combo_dev    = tk.Frame(self)             # combo USB multiple (oculto)
        self.combo_dev      = ttk.Combobox(self.f_combo_dev,
                                            textvariable=self.var_dev)
        self.lbl_carpeta_info = tk.Label(self, text="")
        # Labels de modelo (actualizados por _cargar_modelos_default, no visibles)
        # ya declarados como StringVar en __init__

        # ── Estilo combo ──────────────────────────────────────────────
        style = ttk.Style(); style.theme_use("default")
        style.configure("D.TCombobox", fieldbackground=C_SURFACE2,
                         background=C_SURFACE2, foreground=C_TEXT,
                         selectbackground=C_ACENTO, selectforeground="white",
                         font=(C_MONO, 9))

        # ── Barra de menú ────────────────────────────────────────────────
        menubar = tk.Menu(self, bg=C_SURFACE, fg=C_TEXT,
                          activebackground=C_ACENTO, activeforeground="white",
                          relief="flat", bd=0)
        self.config(menu=menubar)
        menu_ajustes = tk.Menu(menubar, tearoff=0,
                               bg=C_SURFACE, fg=C_TEXT,
                               activebackground=C_ACENTO, activeforeground="white",
                               relief="flat")
        menubar.add_cascade(label="Ajustes", menu=menu_ajustes)
        menu_ajustes.add_command(label="Ajustar parámetros",
                                 command=self._login_ajustes)

        # ── Header ──────────────────────────────────────────────────────
        frame_hdr = tk.Frame(self, bg=C_SURFACE, height=56)
        frame_hdr.pack(fill="x"); frame_hdr.pack_propagate(False)
        tk.Frame(frame_hdr, bg=C_ACENTO, width=4).pack(side="left", fill="y")
        _logo = _cargar_logo()
        if _logo:
            tk.Label(frame_hdr, image=_logo, bg=C_SURFACE
                     ).pack(side="left", padx=14, pady=8)
        else:
            tk.Label(frame_hdr, text="HORSE", bg=C_SURFACE, fg=C_ACENTO,
                     font=(C_MONO, 13, "bold")).pack(side="left", padx=14)
        tk.Frame(frame_hdr, bg=C_BORDER, width=1).pack(side="left", fill="y", pady=8)
        fh = tk.Frame(frame_hdr, bg=C_SURFACE); fh.pack(side="left", padx=14)
        tk.Label(fh, text="ANÁLISIS NVH EN TIEMPO REAL  —  DEMM",
                 bg=C_SURFACE, fg=C_TEXT, font=("Arial", 12, "bold")).pack(anchor="w")
        tk.Label(fh, text="Panel del operador",
                 bg=C_SURFACE, fg=C_TEXT_SUB, font=("Arial", 9)).pack(anchor="w")
        _lm = _hacer_logo_matia(frame_hdr, bg=C_SURFACE)
        _lm.pack(side="right", padx=14, pady=6)
        tk.Frame(self, bg=C_BORDER, height=1).pack(fill="x")

        # ── Cuerpo: columna central estrecha ─────────────────────────
        outer = tk.Frame(self, bg=C_BG)
        outer.pack(fill="both", expand=True)
        # centrar horizontalmente con columnas vacías expansibles
        outer.columnconfigure(0, weight=1)
        outer.columnconfigure(2, weight=1)
        outer.rowconfigure(0, weight=1)

        cuerpo = tk.Frame(outer, bg=C_BG, padx=0, pady=24)
        cuerpo.grid(row=0, column=1, sticky="n")

        def sep(txt):
            tk.Label(cuerpo, text=txt, bg=C_BG, fg=C_ACENTO,
                     font=("Arial", 9, "bold")).pack(anchor="w", pady=(18, 0))
            tk.Frame(cuerpo, bg=C_ACENTO, height=1).pack(fill="x", pady=(2, 8))

        def lbl(txt):
            tk.Label(cuerpo, text=txt, bg=C_BG, fg=C_TEXT,
                     font=("Arial", 10)).pack(anchor="w")

        # ── NOMBRE ───────────────────────────────────────────────────
        sep("OPERADOR")
        lbl("Nombre")
        f_op = tk.Frame(cuerpo, bg=C_BG); f_op.pack(fill="x", pady=(2, 2))
        self.combo_op = ttk.Combobox(f_op, textvariable=self.var_op,
                                      values=self._historial_ops,
                                      style="D.TCombobox",
                                      font=("Arial", 11), width=28)
        self.combo_op.pack(side="left", fill="x", expand=True, ipady=6)
        self.combo_op.bind("<<ComboboxSelected>>", lambda e: self.var_op.set(
            self.combo_op.get()))
        tk.Button(f_op, text="🗑", command=self._borrar_operador,
                  bg=C_MALO, fg="white",
                  activebackground="#7f1d1d", activeforeground="white",
                  relief="flat", bd=0, font=("Arial", 11),
                  cursor="hand2", padx=8
                  ).pack(side="left", padx=(6, 0), ipady=4)
        tk.Label(cuerpo, text="Escribe un nombre nuevo o selecciona uno registrado",
                 bg=C_BG, fg=C_TEXT_DIM, font=("Arial", 7)).pack(anchor="w", pady=(0, 8))

        # ── TURNO ────────────────────────────────────────────────────
        lbl("Turno")
        ttk.Combobox(cuerpo, textvariable=self.var_turno,
                     values=["Mañana", "Tarde", "Noche"],
                     state="readonly", style="D.TCombobox",
                     font=(C_MONO, 10), width=30
                     ).pack(fill="x", ipady=5, pady=(2, 4))

        # Labels ocultos requeridos por otros métodos — no se muestran
        self.lbl_pinon_detectado = tk.Label(self, bg=C_BG, fg=C_BG, font=("Arial", 1))
        self.lbl_pinon_conf      = tk.Label(self, bg=C_BG, fg=C_BG, font=("Arial", 1))
        self.lbl_umbral_estado   = tk.Label(self, bg=C_BG, fg=C_BG, font=("Arial", 1))

        # ── BOTÓN INICIAR ─────────────────────────────────────────────
        tk.Frame(cuerpo, bg=C_BORDER, height=1).pack(fill="x", pady=(14, 14))

        self.btn_ini = tk.Button(cuerpo,
                                  text="▶  INICIAR",
                                  command=self._toggle,
                                  bg=C_ACENTO, fg="white",
                                  activebackground="#3a7fe0",
                                  activeforeground="white",
                                  relief="flat", bd=0,
                                  font=("Arial", 13, "bold"),
                                  cursor="hand2", pady=14, width=32)
        self.btn_ini.pack(fill="x")

        # Label oculto — usado internamente para mensajes de estado
        self.lbl_cargando = tk.Label(self, text="", bg=C_BG,
                                      fg=C_BG, font=("Arial", 1))

        # Indicadores BUENO/MALO — ocultos en esta ventana, usados por _on_ciclo
        self.ind_bueno = tk.Label(self, text="● BUENO", bg=C_SURFACE2,
                                   fg=C_TEXT_DIM, font=(C_MONO, 15, "bold"))
        self.ind_malo  = tk.Label(self, text="✖  MALO",  bg=C_SURFACE2,
                                   fg=C_TEXT_DIM, font=(C_MONO, 15, "bold"))
        # Contadores del turno — ocultos, actualizados por _on_ciclo
        self.lbl_cnt = tk.Label(self, text="BUENO   0\nMALO    0\nREVISAR 0",
                                 bg=C_BG, fg=C_TEXT_SUB, font=(C_MONO, 9),
                                 justify="left")

        tk.Frame(self, bg=C_BORDER, height=1).pack(fill="x", side="bottom")

    def _cargar_modelos_default(self):
        """Carga modelos y umbrales espectrales desde rutas por defecto al iniciar."""
        cargados = []

        # ── Modelo identificador ──────────────────────────────────────
        if os.path.isfile(RUTA_MODELO_ID):
            try:
                self.motor.cargar_id(RUTA_MODELO_ID)
                self.lbl_id.set(f"✓  {os.path.basename(RUTA_MODELO_ID)}")
                cargados.append("ID")
            except Exception as e:
                self.lbl_id.set(f"Error: {e}")

        # ── Modelos NVH ───────────────────────────────────────────────
        if os.path.isfile(RUTA_MODELO_PIMA):
            try:
                key = self.motor.cargar_nvh(RUTA_MODELO_PIMA)
                self.lbl_pima.set(f"\u2713  {os.path.basename(RUTA_MODELO_PIMA)}  [{key}]")
                cargados.append("NVH-PIMA")
            except Exception as e: self.lbl_pima.set(f"Error: {e}")
        if os.path.isfile(RUTA_MODELO_AS):
            try:
                key = self.motor.cargar_nvh(RUTA_MODELO_AS)
                self.lbl_as.set(f"\u2713  {os.path.basename(RUTA_MODELO_AS)}  [{key}]")
                cargados.append("NVH-AS")
            except Exception as e: self.lbl_as.set(f"Error: {e}")

        # ── Umbrales espectrales ──────────────────────────────────────
        umbrales_cargados = []
        for ruta_umb in [RUTA_UMB_AS14, RUTA_UMB_PIMA]:
            if os.path.isfile(ruta_umb):
                try:
                    key, n_arch = self.motor.cargar_umbral_espectro(ruta_umb)
                    umbrales_cargados.append(key)
                except Exception:
                    pass

        if umbrales_cargados:
            msg = f"✓ Umbrales: {', '.join(umbrales_cargados)}"
            self.lbl_umbral_estado.config(text=msg, fg=C_BUENO)
            cargados.append("Umbrales")
        else:
            self.lbl_umbral_estado.config(
                text="Sin umbrales cargados", fg=C_TEXT_DIM)

        if cargados:
            self._st(f"Cargados: {', '.join(cargados)}", C_BUENO)
        else:
            self._st("Modelos no encontrados — carga manual", C_REVISAR)

    def _login_ajustes(self):
        """Ventana de login antes de acceder a los ajustes de parámetros."""
        win = tk.Toplevel(self)
        win.title("Acceso restringido")
        win.geometry("340x220")
        win.resizable(False, False)
        win.configure(bg=C_BG)
        win.grab_set()
        win.transient(self)
        _aplicar_icono(win)
        # Centrar respecto a la ventana principal
        win.update_idletasks()
        x = self.winfo_x() + (self.winfo_width()  - 340) // 2
        y = self.winfo_y() + (self.winfo_height() - 220) // 2
        win.geometry(f"340x220+{x}+{y}")

        tk.Frame(win, bg=C_ACENTO, height=4).pack(fill="x")
        tk.Label(win, text="🔒  Área restringida",
                 bg=C_BG, fg=C_TEXT, font=("Arial", 11, "bold")).pack(pady=(18, 4))
        tk.Label(win, text="Introduce tus credenciales para continuar.",
                 bg=C_BG, fg=C_TEXT_DIM, font=("Arial", 8)).pack(pady=(0, 14))

        f = tk.Frame(win, bg=C_BG); f.pack(fill="x", padx=30)
        tk.Label(f, text="Usuario:", bg=C_BG, fg=C_TEXT,
                 font=("Arial", 9), width=10, anchor="w").grid(row=0, column=0, pady=4)
        var_usr = tk.StringVar()
        tk.Entry(f, textvariable=var_usr, bg=C_SURFACE2, fg=C_TEXT,
                 insertbackground=C_TEXT, font=("Arial", 10),
                 relief="flat", bd=3, width=18).grid(row=0, column=1, pady=4)

        tk.Label(f, text="Contraseña:", bg=C_BG, fg=C_TEXT,
                 font=("Arial", 9), width=10, anchor="w").grid(row=1, column=0, pady=4)
        var_pwd = tk.StringVar()
        tk.Entry(f, textvariable=var_pwd, show="●", bg=C_SURFACE2, fg=C_TEXT,
                 insertbackground=C_TEXT, font=("Arial", 10),
                 relief="flat", bd=3, width=18).grid(row=1, column=1, pady=4)

        lbl_err = tk.Label(win, text="", bg=C_BG, fg=C_MALO, font=("Arial", 8))
        lbl_err.pack(pady=(4, 0))

        def _acceder():
            if var_usr.get().strip() == "admin" and var_pwd.get().strip() == "pe2026":
                win.destroy()
                self._ventana_ajustes()
            else:
                lbl_err.config(text="Usuario o contraseña incorrectos.")
                var_pwd.set("")

        fbtn = tk.Frame(win, bg=C_BG); fbtn.pack(pady=(8, 0))
        tk.Button(fbtn, text="Cancelar", command=win.destroy,
                  bg=C_SURFACE2, fg=C_TEXT_DIM, relief="flat", bd=0,
                  font=("Arial", 9), cursor="hand2", padx=16, pady=6
                  ).pack(side="left", padx=(0, 8))
        tk.Button(fbtn, text="Acceder", command=_acceder,
                  bg=C_ACENTO, fg="white", relief="flat", bd=0,
                  font=("Arial", 9, "bold"), cursor="hand2", padx=16, pady=6
                  ).pack(side="left")

        # Permitir Enter para acceder
        win.bind("<Return>", lambda e: _acceder())

    def _ventana_ajustes(self):
        """Ventana para ajustar umbrales K_ret, K_emp y CF por piñón."""
        # Umbrales originales de referencia (μ+3σ calculados del dataset)
        UMB_ORIG = {
            "PIMA":         {"K_ret": 4.04, "K_emp": 3.89, "CF_ret": 3.27, "CF_emp": 3.34},
            "ARBOL_SEC_14": {"K_ret": 5.23, "K_emp": 4.95, "CF_ret": 3.64, "CF_emp": 3.61},
        }

        win = tk.Toplevel(self)
        win.title("Ajustar parámetros — Umbrales NVH")
        win.geometry("520x540")
        win.resizable(False, False)
        win.configure(bg=C_BG)
        win.grab_set()
        win.transient(self)
        _aplicar_icono(win)
        win.update_idletasks()
        x = self.winfo_x() + (self.winfo_width()  - 520) // 2
        y = self.winfo_y() + (self.winfo_height() - 540) // 2
        win.geometry(f"520x540+{x}+{y}")

        tk.Frame(win, bg=C_ACENTO, height=4).pack(fill="x")
        tk.Label(win, text="⚙  Ajuste de umbrales NVH",
                 bg=C_BG, fg=C_TEXT, font=("Arial", 11, "bold")).pack(pady=(16, 2))
        tk.Label(win, text="Modifica los umbrales de Kurtosis (ret/emp) y CF.p99 por tipo de piñón.",
                 bg=C_BG, fg=C_TEXT_DIM, font=("Arial", 8)).pack(pady=(0, 14))

        vars_umb = {}
        PINONES_UI = [("PIMA", "PIMA  (26 dientes)"),
                      ("ARBOL_SEC_14", "Árbol Secundario  (14 dientes)")]

        for key, nombre in PINONES_UI:
            kr_actual  = UMBRALES[key]["K_ret"]
            ke_actual  = UMBRALES[key]["K_emp"]
            cfr_actual = UMBRALES[key]["CF_ret"]
            cfe_actual = UMBRALES[key]["CF_emp"]
            kr_orig    = UMB_ORIG[key]["K_ret"]
            ke_orig    = UMB_ORIG[key]["K_emp"]
            cfr_orig   = UMB_ORIG[key]["CF_ret"]
            cfe_orig   = UMB_ORIG[key]["CF_emp"]

            tk.Frame(win, bg=C_BORDER, height=1).pack(fill="x", padx=20)
            f = tk.Frame(win, bg=C_SURFACE2, padx=16, pady=10)
            f.pack(fill="x", padx=20, pady=(0, 4))

            tk.Label(f, text=nombre, bg=C_SURFACE2, fg=C_ACENTO,
                     font=("Arial", 10, "bold")).grid(
                     row=0, column=0, columnspan=4, sticky="w", pady=(0, 6))

            # K Retroceso
            tk.Label(f, text="K Retroceso:", bg=C_SURFACE2, fg=C_TEXT,
                     font=("Arial", 9), width=14, anchor="w").grid(row=1, column=0)
            var_kr = tk.StringVar(value=str(kr_actual))
            tk.Entry(f, textvariable=var_kr, bg=C_SURFACE, fg=C_TEXT,
                     insertbackground=C_TEXT, font=(C_MONO, 10),
                     relief="flat", bd=3, width=8).grid(row=1, column=1, padx=(4, 16))
            tk.Label(f, text=f"Original: {kr_orig}",
                     bg=C_SURFACE2, fg=C_TEXT_DIM,
                     font=("Arial", 7)).grid(row=1, column=2, sticky="w")

            # K Empuje
            tk.Label(f, text="K Empuje:", bg=C_SURFACE2, fg=C_TEXT,
                     font=("Arial", 9), width=14, anchor="w").grid(row=2, column=0, pady=(6,0))
            var_ke = tk.StringVar(value=str(ke_actual))
            tk.Entry(f, textvariable=var_ke, bg=C_SURFACE, fg=C_TEXT,
                     insertbackground=C_TEXT, font=(C_MONO, 10),
                     relief="flat", bd=3, width=8).grid(row=2, column=1, padx=(4, 16), pady=(6,0))
            tk.Label(f, text=f"Original: {ke_orig}",
                     bg=C_SURFACE2, fg=C_TEXT_DIM,
                     font=("Arial", 7)).grid(row=2, column=2, sticky="w", pady=(6,0))

            # CF Retroceso
            tk.Label(f, text="CF.p99 Retroceso:", bg=C_SURFACE2, fg=C_TEXT,
                     font=("Arial", 9), width=14, anchor="w").grid(row=3, column=0, pady=(6,0))
            var_cfr = tk.StringVar(value=str(cfr_actual))
            tk.Entry(f, textvariable=var_cfr, bg=C_SURFACE, fg=C_TEXT,
                     insertbackground=C_TEXT, font=(C_MONO, 10),
                     relief="flat", bd=3, width=8).grid(row=3, column=1, padx=(4, 16), pady=(6,0))
            tk.Label(f, text=f"Original: {cfr_orig}",
                     bg=C_SURFACE2, fg=C_TEXT_DIM,
                     font=("Arial", 7)).grid(row=3, column=2, sticky="w", pady=(6,0))

            # CF Empuje
            tk.Label(f, text="CF.p99 Empuje:", bg=C_SURFACE2, fg=C_TEXT,
                     font=("Arial", 9), width=14, anchor="w").grid(row=4, column=0, pady=(6,0))
            var_cfe = tk.StringVar(value=str(cfe_actual))
            tk.Entry(f, textvariable=var_cfe, bg=C_SURFACE, fg=C_TEXT,
                     insertbackground=C_TEXT, font=(C_MONO, 10),
                     relief="flat", bd=3, width=8).grid(row=4, column=1, padx=(4, 16), pady=(6,0))
            tk.Label(f, text=f"Original: {cfe_orig}",
                     bg=C_SURFACE2, fg=C_TEXT_DIM,
                     font=("Arial", 7)).grid(row=4, column=2, sticky="w", pady=(6,0))

            vars_umb[key] = (var_kr, var_ke, var_cfr, var_cfe)

        lbl_msg = tk.Label(win, text="", bg=C_BG, fg=C_BUENO, font=("Arial", 8))
        lbl_msg.pack(pady=(10, 0))

        def _aplicar():
            try:
                nuevos = {}
                for key, (var_kr, var_ke, var_cfr, var_cfe) in vars_umb.items():
                    kr  = float(var_kr.get().strip())
                    ke  = float(var_ke.get().strip())
                    cfr = float(var_cfr.get().strip())
                    cfe = float(var_cfe.get().strip())
                    if kr <= 0 or ke <= 0 or cfr <= 0 or cfe <= 0:
                        raise ValueError("Los valores deben ser positivos.")
                    nuevos[key] = {"K_ret": kr, "K_emp": ke,
                                   "CF_ret": cfr, "CF_emp": cfe}
                for key, vals in nuevos.items():
                    UMBRALES[key]["K_ret"]  = vals["K_ret"]
                    UMBRALES[key]["K_emp"]  = vals["K_emp"]
                    UMBRALES[key]["CF_ret"] = vals["CF_ret"]
                    UMBRALES[key]["CF_emp"] = vals["CF_emp"]
                lbl_msg.config(
                    text="✓ Umbrales actualizados correctamente.", fg=C_BUENO)
                self._st("Umbrales NVH actualizados por el administrador.", C_BUENO)
            except ValueError as e:
                lbl_msg.config(text=f"Error: {e}", fg=C_MALO)

        tk.Frame(win, bg=C_BORDER, height=1).pack(fill="x", padx=20, pady=(12, 8))
        fbtn = tk.Frame(win, bg=C_BG); fbtn.pack()
        tk.Button(fbtn, text="Cancelar", command=win.destroy,
                  bg=C_SURFACE2, fg=C_TEXT_DIM, relief="flat", bd=0,
                  font=("Arial", 9), cursor="hand2", padx=20, pady=8
                  ).pack(side="left", padx=(0, 10))
        tk.Button(fbtn, text="✓  Aplicar", command=_aplicar,
                  bg=C_BUENO, fg="white", relief="flat", bd=0,
                  font=("Arial", 9, "bold"), cursor="hand2", padx=20, pady=8
                  ).pack(side="left")

    def _cargar_historial(self):
        """Carga la lista de operadores guardados desde JSON."""
        try:
            if os.path.isfile(self._ruta_historial):
                with open(self._ruta_historial, "r", encoding="utf-8") as f:
                    import json as _json
                    return _json.load(f)
        except Exception:
            pass
        return []

    def _guardar_historial(self):
        """Guarda la lista de operadores en JSON."""
        try:
            os.makedirs(os.path.dirname(self._ruta_historial), exist_ok=True)
            with open(self._ruta_historial, "w", encoding="utf-8") as f:
                import json as _json
                _json.dump(self._historial_ops, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    def _registrar_operador(self, nombre):
        """Agrega el nombre al historial si no existe y actualiza el combo."""
        nombre = nombre.strip()
        if not nombre:
            return
        if nombre not in self._historial_ops:
            self._historial_ops.insert(0, nombre)
            self._historial_ops = self._historial_ops[:20]  # máx 20
            self._guardar_historial()
            self.combo_op["values"] = self._historial_ops

    def _borrar_operador(self):
        """Borra el nombre seleccionado del historial."""
        nombre = self.var_op.get().strip()
        if not nombre:
            return
        if nombre in self._historial_ops:
            if messagebox.askyesno("Borrar operador",
                                   f"¿Eliminar '{nombre}' del historial?"):
                self._historial_ops.remove(nombre)
                self._guardar_historial()
                self.combo_op["values"] = self._historial_ops
                self.var_op.set("")
        else:
            messagebox.showinfo("No registrado",
                                f"'{nombre}' no está en el historial.")

    def _sel_umbral_espectro(self):
        ruta = filedialog.askopenfilename(
            title="Cargar umbrales espectrales (.pkl)",
            filetypes=[("Pickle", "*.pkl"), ("Todos", "*.*")])
        if not ruta: return
        try:
            key, n_arch = self.motor.cargar_umbral_espectro(ruta)
            msg = f"✓ Umbrales: {key}  ({n_arch} arch.)"
            self.lbl_umbral_estado.config(text=msg, fg=C_BUENO)
            self._st(msg, C_BUENO)
        except Exception as e:
            self.lbl_umbral_estado.config(text=f"Error: {e}", fg=C_MALO)

    def _sel_carpeta_ml(self):
        d = filedialog.askdirectory(title="Carpeta raíz para datos ML")
        if not d: return
        self.var_carpeta.set(d)
        self.motor.carpeta_salida = d
        self.lbl_carpeta_info.config(
            text=f"  {d}  →  PIMA_DATOS_ML/  y  AS_DATOS_ML/",
            fg=C_ACENTO)

    def _listar_devs(self):
        """
        Lógica de selección de micrófono USB:
          - 1 USB  → autoselecciona, muestra solo el label con el nombre.
          - >1 USB → muestra combo para que el operador elija entre los USB.
          - 0 USB  → toma el primer dispositivo de entrada disponible.
          - 0 entradas → error.
        """
        if not SD_OK:
            self._dev_id_auto = None
            if hasattr(self, "lbl_dev_nombre"):
                self.lbl_dev_nombre.config(
                    text="sounddevice no instalado", fg=C_MALO)
            return
        try:
            devs    = sd.query_devices()
            entradas = [(i, d) for i, d in enumerate(devs)
                        if d["max_input_channels"] > 0]
            self._devs = entradas
            usb = [(i, d) for i, d in entradas
                   if "usb" in d["name"].lower()]

            if not entradas:
                # Sin dispositivos en absoluto
                self._dev_id_auto = None
                self.lbl_dev_nombre.config(
                    text="Sin dispositivos de entrada detectados", fg=C_MALO)
                self.f_combo_dev.pack_forget()
                return

            if len(usb) > 1:
                # Mas de un USB: mostrar combo para que el operador elija
                nombres = [f"{i}: {d['name']}" for i, d in usb]
                self.combo_dev["values"] = nombres
                self.combo_dev.current(0)
                self._dev_id_auto = usb[0][0]    # preselección = primero
                self.lbl_dev_nombre.config(
                    text="Varios USB detectados — elige uno:", fg=C_REVISAR)
                self.f_combo_dev.pack(fill="x", pady=(0, 6))
                # Actualizar selección cuando el operador cambie el combo
                self.combo_dev.bind("<<ComboboxSelected>>", self._on_combo_dev)
            else:
                # 1 USB o ninguno (usar primera entrada disponible)
                self.f_combo_dev.pack_forget()
                sel = usb[0] if usb else entradas[0]
                self._dev_id_auto = sel[0]
                nombre = sel[1]["name"]
                prefijo = "\u2713  " if usb else "\u26A0  "
                color   = C_BUENO if usb else C_REVISAR
                self.lbl_dev_nombre.config(text=f"{prefijo}{nombre}", fg=color)

        except Exception as e:
            self._dev_id_auto = None
            if hasattr(self, "lbl_dev_nombre"):
                self.lbl_dev_nombre.config(text=f"Error: {e}", fg=C_MALO)

    def _on_combo_dev(self, _event=None):
        """Callback cuando el operador cambia el combo de dispositivo USB."""
        s = self.var_dev.get().strip()
        try:
            self._dev_id_auto = int(s.split(":")[0])
        except Exception:
            self._dev_id_auto = None

    def _get_dev(self):
        """Retorna el ID del dispositivo autodetectado (USB primero)."""
        return getattr(self, "_dev_id_auto", None)

    def _sel_id(self):
        r = filedialog.askopenfilename(title="Modelo identificador",
                                        filetypes=[("Pickle","*.pkl")])
        if not r: return
        try:
            self.motor.cargar_id(r)
            self.lbl_id.set(f"✓  {os.path.basename(r)}")
            self._st("Modelo identificador cargado", C_BUENO)
        except Exception as e: self.lbl_id.set(f"Error: {e}")

    def _sel_nvh(self, tipo):
        lbl = self.lbl_pima if tipo=="PIMA" else self.lbl_as
        r = filedialog.askopenfilename(title=f"Modelo NVH — {tipo}",
                                        filetypes=[("Pickle","*.pkl")])
        if not r: return
        try:
            key = self.motor.cargar_nvh(r)
            lbl.set(f"✓  {os.path.basename(r)}  [{key}]")
            self._st(f"Modelo {key} cargado", C_BUENO)
        except Exception as e: lbl.set(f"Error: {e}")

    def _toggle(self):
        if not self._activo:
            if not self.var_op.get().strip():
                messagebox.showwarning("Faltan datos","Ingresa el nombre del operador."); return
            # Registrar nombre en historial
            self._registrar_operador(self.var_op.get())
            if not SD_OK:
                self._st("Instala sounddevice: pip install sounddevice", C_MALO); return
            if not self.motor.modelos_nvh:
                messagebox.showwarning("Sin modelos","Carga al menos un modelo NVH."); return
            # Modelo identificador ya no es obligatorio — selección es manual
            # if self.motor.modelo_id is None: ...
            try:
                if self.var_carpeta and self.var_carpeta.get():
                    self.motor.carpeta_salida = self.var_carpeta.get()
                # pinon_manual NO se resetea — el operador lo selecciona
                # manualmente en los cuadros del semáforo.

                # ── Crear ventanas ANTES de iniciar el motor ──────────────
                if self._vs is None or not self._vs.winfo_exists():
                    self._vs = VentanaSemaforo(self)
                    # Conectar selección manual al motor
                    def _cb_manual(pk):
                        self.motor.pinon_manual = pk
                    self._vs._cb_manual = _cb_manual
                if self._vt is None or not self._vt.winfo_exists():
                    self._vt = VentanaTurno(self, self.var_op.get().strip(),
                                             self.var_turno.get())
                if self._vd is None or not self._vd.winfo_exists():
                    self._vd = VentanaDashboard(self)

                # ── Maximizar y enfocar la ventana de Estado ─────────────
                self._vs.state("zoomed")
                self._vs.lift()
                self._vs.focus_force()

                # ── Asignar callbacks ANTES de iniciar ────────────────────
                self.motor.cb_senal = lambda d: (
                    self._vt.push_vivo(d)
                    if self._vt and self._vt.winfo_exists() else None)

                def _on_estado(msg, tipo):
                    col = C_BUENO if tipo=="ok" else C_MALO if tipo=="err" else C_REVISAR
                    self.after(0, lambda m=msg, c=col: self._st(m, c))
                    # Al confirmar engrane → poner semáforo en EN ESPERA
                    if "Engrane confirmado" in msg or "Engrane detectado" in msg:
                        if self._vs and self._vs.winfo_exists():
                            self.after(0, self._vs.iniciar_espera)
                self.motor.cb_estado = _on_estado

                self.motor.cb_ciclo = lambda r: self.after(
                    0, lambda res=r: self._on_ciclo(res))

                # ── Ahora sí iniciar el motor ─────────────────────────────
                self._ciclos_turno = []   # reset al iniciar nuevo turno
                self.motor.iniciar(self._get_dev())
                self._activo = True
                self.btn_ini.config(text="■  DETENER ANÁLISIS",
                                     bg=C_MALO, fg="white",
                                     activebackground=C_MALO)
                self._st(f"Calibrando {int(CALIBRACION_SEG)}s de ruido de fondo...",
                         C_REVISAR)
            except Exception as e: self._st(f"Error al iniciar: {e}", C_MALO)
        else:
            self.motor.detener(); self._activo = False
            self.btn_ini.config(text="▶  INICIAR ANÁLISIS",
                                 bg=C_ACENTO, fg="white",
                                 activebackground=C_ACENTO)
            self._st("Análisis detenido", C_TEXT_DIM)
            self._reset_ind()

    def _on_ciclo(self, res):
        # Guardar ciclo localmente — independiente de _vt
        self._ciclos_turno.append(res)

        etq     = res.get("etiqueta", "?"); self._reset_ind()
        golpe   = res.get("golpe",   False)
        sireneo = res.get("sireneo", False)
        pinon   = res.get("pinon",   "?")
        conf_id = res.get("conf_id", 0.0)

        # ── Actualizar panel de piñón detectado ───────────────────────
        NOMBRES = {"PIMA":         "PIMA  26d",
                   "ARBOL_SEC_14": "ÁRBOL SEC. 14d",
                   "ARBOL_SEC_15": "ÁRBOL SEC. 15d"}
        nombre_pinon = NOMBRES.get(pinon, pinon)
        color_id = C_BUENO if conf_id >= 0.70 else C_REVISAR if conf_id >= 0.50 else C_MALO
        self.lbl_pinon_detectado.config(
            text=nombre_pinon, fg=color_id)
        estado_conf = ("✓ Alta" if conf_id >= 0.70
                       else "⚠ Media" if conf_id >= 0.50
                       else "✗ Baja — verificar piñón")
        self.lbl_pinon_conf.config(
            text=f"Confianza: {conf_id:.0%}  {estado_conf}",
            fg=color_id)
        # Aviso en barra de estado si confianza baja
        if conf_id < 0.70:
            self._st(f"⚠ Confianza baja ({conf_id:.0%}) — piñón detectado: {nombre_pinon}",
                     C_REVISAR)

        # ── Actualizar semáforo ────────────────────────────────────────
        if self._vs and self._vs.winfo_exists():
            self._vs.actualizar(
                golpe, sireneo,
                pinon         = pinon,
                zona_golpe    = res.get("zona_golpe", ""),
                armon_sireneo = res.get("armon_sireneo", []))

        # ── Indicadores panel operador ─────────────────────────────────
        if sireneo:
            # DESECHAR — sireneo es irrecuperable
            self.ind_malo.config(bg=C_MALO, fg="white", text="DESECHAR")
            self.after(2500, self._reset_ind)
        elif golpe:
            # REPARAR — golpe puede limarse
            self.ind_malo.config(bg=C_REVISAR, fg="white", text="REPARAR")
            self.after(2500, self._reset_ind)
        else:
            # BUENO
            self.ind_bueno.config(bg=C_BUENO, fg="white", text="BUENO")
            self.after(2500, self._reset_ind)
        # Actualizar tabla
        if self._vt and self._vt.winfo_exists():
            self._vt.agregar_ciclo(res)
        # Actualizar dashboard en tiempo real
        if self._vd and self._vd.winfo_exists():
            self.after(50, lambda r=res: self._vd.agregar_ciclo_dashboard(r))
            nb   = sum(1 for c in self._vt.ciclos
                       if not c.get("golpe") and not c.get("sireneo"))
            nrep = sum(1 for c in self._vt.ciclos
                       if c.get("golpe") and not c.get("sireneo"))
            ndes = sum(1 for c in self._vt.ciclos if c.get("sireneo"))
            self.lbl_cnt.config(
                text=f"BUENO    {nb}\nREPARAR  {nrep}\nDESECHAR {ndes}")

    def _reset_ind(self):
        self.ind_bueno.config(bg=C_SURFACE2, fg=C_TEXT_DIM)
        self.ind_malo.config( bg=C_SURFACE2, fg=C_TEXT_DIM)

    def _st(self, msg, color=C_TEXT_DIM):
        self.lbl_cargando.config(text=msg, fg=color)

    def run(self):
        def _cerrar():
            self.motor.detener()

            ciclos   = self._ciclos_turno
            operador = self.var_op.get().strip() or "Operador"
            turno    = self.var_turno.get()

            if not ciclos:
                self.destroy()
                return

            # Nombre propuesto
            fecha_hora  = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            nombre_prop = (f"turno_{operador.replace(' ','_')}_"
                           f"{turno}_{fecha_hora}.xlsx")

            # Carpeta inicial sugerida
            fecha_hoy   = datetime.datetime.now().strftime("%d-%m-%Y")
            carpeta_ini = os.path.join(
                self.motor.carpeta_salida, f"DIV-33_{fecha_hoy}")
            try:
                os.makedirs(carpeta_ini, exist_ok=True)
            except Exception:
                carpeta_ini = os.path.expanduser("~")

            # Diálogo nativo de Windows
            ruta = filedialog.asksaveasfilename(
                title="Guardar tabla del turno — MAT IA",
                initialdir=carpeta_ini,
                initialfile=nombre_prop,
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")])

            if not ruta:
                # Usuario canceló → preguntar si quiere cerrar sin guardar
                if messagebox.askyesno(
                        "Cerrar sin guardar",
                        f"¿Cerrar sin guardar?\n\n"
                        f"Se perderán {len(ciclos)} ciclos registrados."):
                    self.destroy()
                return

            try:
                exportar_excel(ciclos, operador, turno, ruta)
                messagebox.showinfo(
                    "MAT IA — Guardado exitoso",
                    f"Tabla guardada correctamente:\n\n{os.path.basename(ruta)}\n\n"
                    f"Ubicación:\n{os.path.dirname(ruta)}")
                self.destroy()
            except Exception as e:
                messagebox.showerror(
                    "Error al guardar",
                    f"No se pudo guardar el archivo:\n\n{e}")

        self.protocol("WM_DELETE_WINDOW", _cerrar)
        self.after(100, lambda: self.wm_state("zoomed"))
        self.mainloop()



def _aplicar_icono(ventana):
    """
    Aplica matia.ico en barra de título Y barra de tareas de Windows.

    En Windows, iconbitmap() sobre una ventana con overrideredirect(True)
    no afecta la barra de tareas. La solución es:
      1. wm_iconbitmap(default=ruta) — establece el ícono por defecto
         para TODAS las ventanas Toplevel de esta instancia Tk, incluyendo
         la entrada en la barra de tareas.
      2. La ruta debe ser absoluta y con barras invertidas en Windows.
      3. iconphoto(True, ...) como fallback multiplataforma.
    """
    ruta_ico  = _resource_path("matia.ico")
    ruta_png  = _resource_path("matia_logo.png")
    try:
        if os.path.exists(ruta_ico):
            # Ruta con barras invertidas para Windows
            ruta_win = ruta_ico.replace("/", "\\")
            try:
                # wm_iconbitmap con default aplica a toda la jerarquía Tk
                # incluida la entrada de la barra de tareas
                ventana.wm_iconbitmap(default=ruta_ico)
            except Exception:
                pass
            try:
                ventana.iconbitmap(ruta_ico)
            except Exception:
                pass
            # Forzar actualización del ícono en la barra de tareas
            # usando after() para aplicar tras el primer render
            try:
                ventana.after(50, lambda: ventana.iconbitmap(ruta_ico))
            except Exception:
                pass
            return
        # Fallback PNG con iconphoto (multiplataforma)
        if os.path.exists(ruta_png):
            from PIL import Image, ImageTk
            img = Image.open(ruta_png).convert("RGBA")
            # Crear versión cuadrada para el ícono
            s   = 64
            canvas = Image.new("RGBA", (s, s), (0,0,0,0))
            img.thumbnail((s, s), Image.LANCZOS)
            canvas.paste(img, ((s-img.width)//2, (s-img.height)//2), img)
            photo = ImageTk.PhotoImage(canvas)
            ventana.iconphoto(True, photo)
            ventana._icon_ref = photo
    except Exception:
        pass

class SplashScreen(tk.Tk):
    """
    Pantalla de bienvenida MatIA — 5 segundos.

    Engranaje: dientes como rectángulos rotados (igual al diseño original),
               aparece girando con fade-in.
    Texto "Mat" e "IA": cargado desde PNG para fidelidad tipográfica,
                          animado por separado.
    Secuencia:
      0  –  600ms : engranaje gira y aparece (dientes rectangulares)
      600 – 1100ms : "Mat" desliza +50px→0 desde izquierda con fade-in
      1100 – 1600ms: "IA"   desliza -50px→0 desde derecha  con fade-in
      1600 – 1900ms: línea acento se traza de izq a der
      1900 – 2200ms: tagline fade-in
      2200 – 4600ms: logo completo visible (pausa)
      4600 – 5000ms: fade-out y lanzar app
    """

    W, H     = 600, 280
    TOTAL_MS = 4600

    # Colores
    BG      = "#f0f2f5"
    C_DARK  = "#1a1d27"
    C_BLUE  = "#1a5fa8"
    C_MID   = "#4a5068"
    C_DIM   = "#8a90a8"

    def __init__(self, on_cerrar):
        super().__init__()
        self._on_cerrar = on_cerrar
        self._frame     = 0
        self._alive     = True
        self._tk_refs   = {}   # mantener referencias PhotoImage
        self._text_imgs = {}   # capas PIL de texto
        self._pil_ok    = False

        # Aplicar ícono ANTES de overrideredirect para que
        # Windows lo registre en la barra de tareas
        _aplicar_icono(self)
        self.overrideredirect(True)
        self.configure(bg=self.BG)
        self.resizable(False, False)
        self.lift()
        self.attributes("-topmost", True)

        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        self.geometry(f"{self.W}x{self.H}+"
                      f"{(sw-self.W)//2}+{(sh-self.H)//2}")

        self._cargar_texto_png()
        self._build()
        self.after(16, self._tick)

    # ─────────────────────────────────────────────────────────────────────
    # Cargar capas de texto desde PNG (tipografía fiel)
    # ─────────────────────────────────────────────────────────────────────
    def _cargar_texto_png(self):
        try:
            from PIL import Image, ImageDraw, ImageFont
            import os

            SCALE = 3
            sz    = 68 * SCALE
            f_bold = None
            f_reg  = None
            for p in ["/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
                      "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
                      "C:/Windows/Fonts/arialbd.ttf",
                      "C:/Windows/Fonts/Arial Bold.ttf"]:
                if os.path.exists(p):
                    try:
                        f_bold = ImageFont.truetype(p, sz)
                        break
                    except Exception:
                        pass
            for p in ["/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
                      "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
                      "C:/Windows/Fonts/arial.ttf",
                      "C:/Windows/Fonts/Arial.ttf"]:
                if os.path.exists(p):
                    try:
                        f_reg = ImageFont.truetype(p, sz)
                        break
                    except Exception:
                        pass

            if not f_bold or not f_reg:
                return

            # Medir texto
            tmp = Image.new("RGBA", (3000, 400), (0,0,0,0))
            dt  = ImageDraw.Draw(tmp)
            pad = 8 * SCALE

            dt.text((pad, 200), "Mat", font=f_bold, fill=(26,29,39,255), anchor="ls")
            bb_m = dt.textbbox((pad, 200), "Mat", font=f_bold, anchor="ls")

            x_ia = bb_m[2] + 6 * SCALE
            dt.text((x_ia, 200), "IA", font=f_reg, fill=(26,95,168,255), anchor="ls")
            bb_i = dt.textbbox((x_ia, 200), "IA", font=f_reg, anchor="ls")

            h_text = bb_m[3] - bb_m[1] + pad * 2

            # Capa "Mat" (RGBA fondo transparente)
            img_m = Image.new("RGBA",
                              (bb_m[2] - bb_m[0] + pad*2, h_text), (0,0,0,0))
            dm = ImageDraw.Draw(img_m)
            dm.text((pad - bb_m[0], pad - bb_m[1]),
                    "Mat", font=f_bold, fill=(26,29,39,255))
            self._text_imgs["math"] = img_m.resize(
                ((bb_m[2]-bb_m[0]+pad*2)//SCALE, h_text//SCALE), Image.LANCZOS)

            # Capa "IA"
            img_i = Image.new("RGBA",
                              (bb_i[2] - bb_i[0] + pad*2, h_text), (0,0,0,0))
            di = ImageDraw.Draw(img_i)
            di.text((pad, pad - bb_m[1]),
                    "IA", font=f_reg, fill=(26,95,168,255))
            self._text_imgs["ia"] = img_i.resize(
                ((bb_i[2]-bb_i[0]+pad*2)//SCALE, h_text//SCALE), Image.LANCZOS)

            # Guardar medidas en coordenadas 1x
            self._math_w = (bb_m[2]-bb_m[0]) // SCALE
            self._ia_w   = (bb_i[2]-bb_i[0]) // SCALE
            self._text_h = h_text // SCALE
            self._pil_ok = True

        except Exception:
            self._pil_ok = False

    # ─────────────────────────────────────────────────────────────────────
    # Build
    # ─────────────────────────────────────────────────────────────────────
    def _build(self):
        self.configure(highlightbackground="#c8cdd8", highlightthickness=1)
        tk.Frame(self, bg=C_ACENTO, height=4).pack(fill="x")
        self.cv = tk.Canvas(self, width=self.W, height=self.H,
                            bg=self.BG, highlightthickness=0)
        self.cv.pack(fill="both", expand=True)
        # Barra progreso
        bar_bg = tk.Frame(self, bg="#dde0e8", height=3)
        bar_bg.pack(fill="x")
        self._bar = tk.Canvas(bar_bg, height=3, bg="#dde0e8",
                              highlightthickness=0)
        self._bar.pack(fill="x")
        self._bar_rect = self._bar.create_rectangle(
            0, 0, 0, 3, fill=C_ACENTO, outline="")

    # ─────────────────────────────────────────────────────────────────────
    # Loop animación
    # ─────────────────────────────────────────────────────────────────────
    def _tick(self):
        if not self._alive or not self.winfo_exists():
            return
        t = self._frame * 16
        self._frame += 1
        self._dibujar(t)
        bw = int(self.W * min(t, self.TOTAL_MS) / self.TOTAL_MS)
        self._bar.coords(self._bar_rect, 0, 0, bw, 3)
        if t < self.TOTAL_MS:
            self.after(16, self._tick)
        else:
            self.after(300, lambda: self._fade_out(0))

    @staticmethod
    def _ease(t, t0, t1):
        if t <= t0: return 0.0
        if t >= t1: return 1.0
        p = (t - t0) / (t1 - t0)
        return p * p * (3 - 2 * p)

    @staticmethod
    def _blend_color(hex_col, alpha, bg=(240,242,245)):
        r = int(hex_col[1:3],16)
        g = int(hex_col[3:5],16)
        b = int(hex_col[5:7],16)
        r = int(bg[0] + (r-bg[0])*alpha)
        g = int(bg[1] + (g-bg[1])*alpha)
        b = int(bg[2] + (b-bg[2])*alpha)
        return f"#{r:02x}{g:02x}{b:02x}"

    # ─────────────────────────────────────────────────────────────────────
    # Dibujo principal
    # ─────────────────────────────────────────────────────────────────────
    def _dibujar(self, t):
        import math
        cv = self.cv
        cv.delete("all")
        self._tk_refs.clear()

        W, H = self.W, self.H
        # Centro del engranaje
        gx, gy = 100, H // 2

        # ── ENGRANAJE ────────────────────────────────────────────────────
        # Dientes: 8 rectángulos de 14x14 rotados cada 45°
        # idéntico al diseño SVG original
        p_gear = self._ease(t, 0, 600)
        rot    = (t / 16) * 1.5   # grados acumulados

        col_dark = self._blend_color(self.C_DARK, p_gear)
        col_blue = self._blend_color(self.C_BLUE, p_gear)
        col_mid  = self._blend_color(self.C_MID,  p_gear)

        # Dientes (rectángulos rotados)
        for i in range(8):
            angle = math.radians(i * 45 + rot)
            # 4 vértices del rect 14×14 centrado en (0,-46) antes de rotar
            local = [(-7,-53),(7,-53),(7,-39),(-7,-39)]
            pts   = []
            for lx, ly in local:
                rx = lx*math.cos(angle) - ly*math.sin(angle)
                ry = lx*math.sin(angle) + ly*math.cos(angle)
                pts.extend([gx + rx, gy + ry])
            cv.create_polygon(pts, fill=col_blue, outline="", smooth=False)

        # Cuerpo oscuro
        r_body = 34
        cv.create_oval(gx-r_body, gy-r_body, gx+r_body, gy+r_body,
                       fill=col_dark, outline=col_blue, width=2)

        # Nodo central
        r_node = 7
        cv.create_oval(gx-r_node, gy-r_node, gx+r_node, gy+r_node,
                       fill=col_blue, outline="")

        # Líneas radiales + satélites (rotan con el engranaje)
        for ad in [0, 90, 180, 270]:
            a  = math.radians(ad + rot)
            x1 = gx + r_node * math.cos(a)
            y1 = gy + r_node * math.sin(a)
            x2 = gx + (r_body - 4) * math.cos(a)
            y2 = gy + (r_body - 4) * math.sin(a)
            cv.create_line(x1, y1, x2, y2, fill=col_mid, width=2)
            sr = 4
            cv.create_oval(x2-sr, y2-sr, x2+sr, y2+sr,
                           fill=col_blue, outline="")

        # ── TEXTO ────────────────────────────────────────────────────────
        tx_base = 160   # x donde empieza "Mat"
        ty_base = H//2 - 30

        if self._pil_ok:
            from PIL import Image, ImageTk

            # "Mat" — desliza desde izquierda
            p_math = self._ease(t, 600, 1100)
            if p_math > 0:
                offset = int((1 - p_math) * 50)
                img_m  = self._text_imgs["math"].copy().convert("RGBA")
                # Aplicar alpha mezclando con BG
                r_,g_,b_,a_ = img_m.split()
                new_a = a_.point(lambda x: int(x * p_math))
                img_m.putalpha(new_a)
                bg_m = Image.new("RGBA", img_m.size, (240,242,245,255))
                bg_m.paste(img_m, mask=img_m.split()[3])
                tk_m = ImageTk.PhotoImage(bg_m.convert("RGB"))
                cv.create_image(tx_base - offset, ty_base,
                                anchor="nw", image=tk_m)
                self._tk_refs["math"] = tk_m

            # "IA" — desliza desde derecha
            p_ia = self._ease(t, 1100, 1600)
            if p_ia > 0:
                offset = int((1 - p_ia) * 50)
                img_i  = self._text_imgs["ia"].copy().convert("RGBA")
                r_,g_,b_,a_ = img_i.split()
                new_a = a_.point(lambda x: int(x * p_ia))
                img_i.putalpha(new_a)
                bg_i = Image.new("RGBA", img_i.size, (240,242,245,255))
                bg_i.paste(img_i, mask=img_i.split()[3])
                tk_i = ImageTk.PhotoImage(bg_i.convert("RGB"))
                x_ia = tx_base + self._math_w
                cv.create_image(x_ia + offset, ty_base,
                                anchor="nw", image=tk_i)
                self._tk_refs["ia"] = tk_i
        else:
            # Fallback texto tkinter
            p_math = self._ease(t, 600, 1100)
            p_ia   = self._ease(t, 1100, 1600)
            if p_math > 0:
                cv.create_text(tx_base - int((1-p_math)*50), ty_base + 40,
                               text="Mat", font=("Arial", 44, "bold"),
                               fill=self._blend_color(self.C_DARK, p_math),
                               anchor="w")
            if p_ia > 0:
                cv.create_text(tx_base + 160 + int((1-p_ia)*50), ty_base + 40,
                               text="IA", font=("Arial", 44),
                               fill=self._blend_color(self.C_BLUE, p_ia),
                               anchor="w")

        # ── LÍNEA ACENTO ─────────────────────────────────────────────────
        p_line = self._ease(t, 1600, 1900)
        if p_line > 0:
            line_x1 = tx_base
            line_x2 = tx_base + int((self._math_w + self._ia_w + 8) * p_line)
            line_y  = ty_base + self._text_h + 4 if self._pil_ok else ty_base + 62
            cv.create_rectangle(line_x1, line_y, line_x2, line_y + 3,
                                fill=C_ACENTO, outline="")

        # ── TAGLINE ───────────────────────────────────────────────────────
        p_tag = self._ease(t, 1900, 2200)
        if p_tag > 0:
            tag_y = ty_base + self._text_h + 14 if self._pil_ok else ty_base + 76
            cv.create_text(tx_base, tag_y,
                           text="ANÁLISIS NVH  ·  DEMM",
                           font=("Consolas", 11),
                           fill=self._blend_color(self.C_MID, p_tag),
                           anchor="w")

        # ── v1.0 ──────────────────────────────────────────────────────────
        p_ver = self._ease(t, 2200, 2500)
        if p_ver > 0:
            cv.create_text(W - 16, H - 14, text="v1.0",
                           font=("Consolas", 9),
                           fill=self._blend_color(self.C_DIM, p_ver),
                           anchor="e")

    # ─────────────────────────────────────────────────────────────────────
    # Fade-out
    # ─────────────────────────────────────────────────────────────────────
    def _fade_out(self, step):
        if not self.winfo_exists():
            self._cerrar(); return
        try:
            self.attributes("-alpha", max(0.0, 1.0 - step / 12))
        except Exception:
            pass
        if step < 12:
            self.after(30, lambda: self._fade_out(step + 1))
        else:
            self._cerrar()

    def _cerrar(self):
        self._alive = False
        if self.winfo_exists():
            self.destroy()
        self._on_cerrar()


if __name__ == "__main__":
    if not SD_OK: print("ADVERTENCIA: pip install sounddevice")

    # ── Arranque: el splash ES la ventana raíz (tk.Tk) ───────────────────
    # Al cerrarse el splash se destruye y VentanaOperador (también tk.Tk)
    # arranca su propio mainloop limpio.
    def _lanzar_app():
        """Lanzado por SplashScreen._cerrar() — crea la app principal."""
        app = VentanaOperador()
        app.protocol("WM_DELETE_WINDOW",
                     lambda: (app.motor.detener(), app.destroy()))
        app.mainloop()

    splash = SplashScreen(on_cerrar=_lanzar_app)
    splash.mainloop()
